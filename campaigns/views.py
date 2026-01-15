from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from .models import (
    Industry, Client, Campaign, AdGroup, Ad, Keyword, PerformanceDataImport,
    HistoricalCampaignPerformance, HistoricalKeywordPerformance,
    NegativeKeywordList, NegativeKeyword, CampaignNegativeKeywordList, NegativeKeywordUpload,
    GeographicRegion, DanishCity, GeographicRegionUpload,
    IndustryService, ServiceKeyword, IndustryHeadline,
    PostalCode
)

# Import geographic regions views
from .geographic_views import (
    geographic_regions_manager, create_geographic_region_ajax, add_danish_city_ajax,
    delete_danish_city_ajax, update_danish_city_ajax, delete_geographic_region_ajax,
    edit_geographic_region_ajax, download_danish_cities_template, import_danish_cities_excel,
    analyze_excel_import_cities, execute_excel_import_cities, suggest_postal_code_ajax,
    generate_negative_city_list, get_negative_city_count
)
from usps.models import USPTemplate, ClientUSP, USPMainCategory
from crawler.tasks import crawl_client_website
from crawler.models import ExtractedUSP
from .data_import import GoogleAdsDataImporter
from .google_ads_export import GoogleAdsEditorExporter, export_simple_csv_format, create_basic_campaign_template
from .geo_export import GeoMarketingExporter, GeoCampaignManager, create_demo_geo_template
from .geo_utils import DanishSlugGenerator, GeoKeywordGenerator, validate_geo_data
from .models import GeoTemplate, GeoKeyword, GeoExport
import json
import os
import tempfile


def campaign_builder(request):
    """Main campaign builder interface"""
    if request.method == 'POST':
        return create_campaign(request)
    
    # Get data for form
    industries = Industry.objects.all()
    usp_templates = USPTemplate.objects.filter(is_active=True)[:10]  # Limit for demo
    
    context = {
        'industries': industries,
        'usp_templates': usp_templates,
    }
    
    return render(request, 'campaigns/builder.html', context)


def create_campaign(request):
    """Process campaign creation form"""
    try:
        # Extract form data
        client_name = request.POST.get('client_name')
        industry_id = request.POST.get('industry')
        website_url = request.POST.get('website_url')
        description = request.POST.get('description', '')
        
        campaign_name = request.POST.get('campaign_name')
        campaign_type = request.POST.get('campaign_type')
        daily_budget = request.POST.get('daily_budget')
        target_location = request.POST.get('target_location')
        bidding_strategy = request.POST.get('bidding_strategy')
        target_cpa = request.POST.get('target_cpa')
        
        selected_usps = request.POST.getlist('selected_usps')
        custom_usps = [
            request.POST.get(f'custom_usp_{i}') 
            for i in range(1, 10) 
            if request.POST.get(f'custom_usp_{i}')
        ]
        
        # Validation
        if not all([client_name, industry_id, website_url, campaign_name, daily_budget]):
            messages.error(request, 'Alle påkrævede felter skal udfyldes')
            return redirect('campaign_builder')
        
        # Create or get industry
        industry = Industry.objects.get(id=industry_id)
        
        # Create or get client (for demo - in production you'd want proper user management)
        client, created = Client.objects.get_or_create(
            name=client_name,
            website_url=website_url,
            defaults={
                'industry': industry,
                'description': description,
                'created_by': request.user if request.user.is_authenticated else None
            }
        )
        
        # Create campaign
        campaign = Campaign.objects.create(
            name=campaign_name,
            client=client,
            campaign_type=campaign_type,
            budget_daily=daily_budget,
            target_location=target_location,
            status='draft'
        )
        
        # Add selected USPs
        for usp_id in selected_usps:
            usp_template = USPTemplate.objects.get(id=usp_id)
            ClientUSP.objects.create(
                client=client,
                usp_template=usp_template,
                custom_text=usp_template.text,
                is_selected=True
            )
        
        # Add custom USPs
        for custom_usp in custom_usps:
            if custom_usp.strip():
                ClientUSP.objects.create(
                    client=client,
                    custom_text=custom_usp.strip(),
                    is_selected=True
                )
        
        # Start website crawling in background
        if request.POST.get('generate_keywords'):
            crawl_client_website.delay(client.id, max_pages=10)
        
        # Generate basic ad group and keywords (simplified for demo)
        if request.POST.get('generate_ads'):
            create_basic_campaign_structure(campaign, client)
        
        messages.success(request, f'Kampagne "{campaign_name}" er oprettet! Website crawling er startet i baggrunden.')
        return redirect('campaign_success', campaign_id=campaign.id)
        
    except Exception as e:
        messages.error(request, f'Fejl ved oprettelse af kampagne: {str(e)}')
        return redirect('campaign_builder')


def create_basic_campaign_structure(campaign, client):
    """Create basic campaign structure with ad groups, keywords and ads"""
    
    # Get client USPs for ad creation
    client_usps = ClientUSP.objects.filter(client=client, is_selected=True)[:3]
    
    # Create main ad group
    ad_group = AdGroup.objects.create(
        name=f"{campaign.name} - Hovedgruppe",
        campaign=campaign,
        default_cpc=15.00,  # 15 DKK default CPC
        priority_score=100
    )
    
    # Create basic keywords based on industry and client name
    industry_keywords = [
        client.industry.name.lower(),
        f"{client.industry.name.lower()} {client.target_location}",
        f"{client.name.split()[0].lower()} {client.industry.name.lower()}",
    ]
    
    for keyword_text in industry_keywords:
        if keyword_text:
            Keyword.objects.create(
                text=keyword_text,
                ad_group=ad_group,
                match_type='phrase',
                max_cpc=15.00
            )
    
    # Create ads using USPs
    usp_texts = [usp.custom_text for usp in client_usps]
    
    # Create first ad
    headline_1 = f"{client.industry.name} i {campaign.target_location}"[:30]
    headline_2 = usp_texts[0][:30] if usp_texts else "Professionel service"
    headline_3 = usp_texts[1][:30] if len(usp_texts) > 1 else ""
    
    description_1 = f"Få hjælp fra {client.name}. " + (usp_texts[2] if len(usp_texts) > 2 else "Ring i dag!")
    description_1 = description_1[:90]
    
    Ad.objects.create(
        ad_group=ad_group,
        headline_1=headline_1,
        headline_2=headline_2,
        headline_3=headline_3,
        description_1=description_1,
        description_2="Ring i dag for gratis tilbud!"[:90],
        final_url=client.website_url,
        display_path_1=client.industry.name.lower()[:15],
        display_path_2=campaign.target_location.split(',')[0][:15]
    )


@csrf_exempt
def crawl_website_ajax(request):
    """AJAX endpoint for website crawling"""
    if request.method == 'POST':
        website_url = request.POST.get('website_url')
        
        if not website_url:
            return JsonResponse({'error': 'Ingen URL angivet'})
        
        try:
            # For demo purposes, we'll create a temporary client
            from django.contrib.auth.models import User
            from crawler.services import WebsiteCrawler
            
            # Get or create default industry and user
            industry, _ = Industry.objects.get_or_create(
                name='Test Industry',
                defaults={'description': 'Temporary industry for crawling'}
            )
            
            user = User.objects.first()
            if not user:
                user = User.objects.create_user('demo', 'demo@example.com', 'demo123')
            
            # Create temporary client
            temp_client = Client.objects.create(
                name=f'Temp - {website_url}',
                website_url=website_url,
                industry=industry,
                created_by=user
            )
            
            # Start crawling
            crawler = WebsiteCrawler(temp_client, max_pages=3, delay=0.5)
            crawl_session = crawler.crawl_website()
            
            # Get extracted USPs
            extracted_usps = ExtractedUSP.objects.filter(
                web_page__crawl_session=crawl_session
            )[:10]
            
            usp_data = []
            for usp in extracted_usps:
                usp_data.append({
                    'text': usp.text,
                    'confidence': usp.confidence_score,
                    'position': usp.position_on_page
                })
            
            return JsonResponse({
                'success': True,
                'usps_found': len(usp_data),
                'usps': usp_data,
                'pages_crawled': crawl_session.pages_crawled
            })
            
        except Exception as e:
            return JsonResponse({'error': f'Fejl ved crawling: {str(e)}'})
    
    return JsonResponse({'error': 'Kun POST requests tilladt'})


def campaign_success(request, campaign_id):
    """Campaign creation success page"""
    try:
        campaign = Campaign.objects.get(id=campaign_id)
        ad_groups = AdGroup.objects.filter(campaign=campaign)
        keywords = Keyword.objects.filter(ad_group__campaign=campaign)
        ads = Ad.objects.filter(ad_group__campaign=campaign)
        
        context = {
            'campaign': campaign,
            'ad_groups_count': ad_groups.count(),
            'keywords_count': keywords.count(),
            'ads_count': ads.count(),
            'ad_groups': ad_groups,
            'keywords': keywords[:10],  # Show first 10
            'ads': ads
        }
        
        return render(request, 'campaigns/success.html', context)
        
    except Campaign.DoesNotExist:
        messages.error(request, 'Kampagne ikke fundet')
        return redirect('campaign_builder')


def home(request):
    """Home page with campaign options"""
    return render(request, 'campaigns/home.html')


def data_import_dashboard(request):
    """Dashboard til at importere Google Ads performance data"""
    
    # Get recent imports
    recent_imports = PerformanceDataImport.objects.filter(
        imported_by=request.user if request.user.is_authenticated else None
    ).order_by('-imported_at')[:10]
    
    # Get statistics
    total_campaigns = HistoricalCampaignPerformance.objects.count()
    total_keywords = HistoricalKeywordPerformance.objects.count()
    
    context = {
        'recent_imports': recent_imports,
        'total_campaigns': total_campaigns,
        'total_keywords': total_keywords,
    }
    
    return render(request, 'campaigns/data_import.html', context)


@csrf_exempt
def import_performance_data(request):
    """AJAX endpoint til at importere performance data"""
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Kun POST requests tilladt'})
    
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'Ingen fil uploaded'})
    
    file = request.FILES['file']
    import_type = request.POST.get('import_type', 'campaign')
    
    # Validate file type
    allowed_extensions = ['.csv', '.xlsx', '.xls']
    file_extension = os.path.splitext(file.name)[1].lower()
    
    if file_extension not in allowed_extensions:
        return JsonResponse({'error': 'Kun CSV og Excel filer er tilladt'})
    
    try:
        # Save file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as temp_file:
            for chunk in file.chunks():
                temp_file.write(chunk)
            temp_file_path = temp_file.name
        
        # Create importer
        user = request.user if request.user.is_authenticated else None
        if not user:
            # Create a default user for demo
            from django.contrib.auth.models import User
            user, created = User.objects.get_or_create(
                username='demo',
                defaults={'email': 'demo@example.com'}
            )
        
        importer = GoogleAdsDataImporter(user)
        
        # Import based on type
        if import_type == 'campaign':
            result = importer.import_campaign_performance(temp_file_path, file.name)
        elif import_type == 'keyword':
            result = importer.import_keyword_performance(temp_file_path, file.name)
        elif import_type == 'account_structure':
            result = importer.import_account_structure(temp_file_path, file.name)
        else:
            result = {'success': False, 'error': 'Ukendt import type'}
        
        # Clean up temp file
        os.unlink(temp_file_path)
        
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({'error': f'Import fejl: {str(e)}'})


def performance_data_overview(request):
    """Oversigt over importeret performance data"""
    
    # Campaign performance summary
    campaigns = HistoricalCampaignPerformance.objects.select_related('import_batch').order_by('-created_at')[:50]
    
    # Keywords by industry
    keywords_by_industry = HistoricalKeywordPerformance.objects.values(
        'industry_category'
    ).distinct().order_by('industry_category')
    
    # Top performing campaigns (lowest cost per conversion)
    top_campaigns = HistoricalCampaignPerformance.objects.filter(
        cost_per_conversion__isnull=False,
        conversions__gte=5  # Minimum 5 conversions
    ).order_by('cost_per_conversion')[:20]
    
    context = {
        'campaigns': campaigns,
        'keywords_by_industry': keywords_by_industry,
        'top_campaigns': top_campaigns,
    }
    
    return render(request, 'campaigns/performance_overview.html', context)


@csrf_exempt
def analyze_patterns_ajax(request):
    """AJAX endpoint til at køre pattern analyse"""
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Kun POST requests tilladt'})
    
    try:
        from .pattern_analyzer import PerformancePatternAnalyzer
        
        analyzer = PerformancePatternAnalyzer()
        results = analyzer.analyze_all_patterns()
        
        return JsonResponse({
            'success': True,
            'results': results
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


def pattern_results(request):
    """Vis resultater af pattern analyse"""
    
    # Get all patterns
    from .models import IndustryPerformancePattern
    
    patterns = IndustryPerformancePattern.objects.all().order_by('industry_name', 'pattern_type')
    
    # Group by industry
    patterns_by_industry = {}
    for pattern in patterns:
        if pattern.industry_name not in patterns_by_industry:
            patterns_by_industry[pattern.industry_name] = {}
        patterns_by_industry[pattern.industry_name][pattern.pattern_type] = pattern
    
    context = {
        'patterns_by_industry': patterns_by_industry,
        'total_patterns': patterns.count(),
        'industries_count': len(patterns_by_industry)
    }
    
    return render(request, 'campaigns/pattern_results.html', context)


def export_campaign_to_google_ads(request, campaign_id):
    """Eksporter kampagne til Google Ads Editor Excel format"""
    try:
        exporter = GoogleAdsEditorExporter(campaign_id)
        return exporter.export_campaign_csv()
    except Campaign.DoesNotExist:
        messages.error(request, 'Kampagne ikke fundet')
        return redirect('campaign_builder')
    except Exception as e:
        messages.error(request, f'Fejl ved eksport: {str(e)}')
        return redirect('campaign_success', campaign_id=campaign_id)


def export_campaign_csv(request, campaign_id):
    """Eksporter kampagne til simpel CSV format"""
    try:
        return export_simple_csv_format(campaign_id)
    except Campaign.DoesNotExist:
        messages.error(request, 'Kampagne ikke fundet')
        return redirect('campaign_builder')
    except Exception as e:
        messages.error(request, f'Fejl ved CSV eksport: {str(e)}')
        return redirect('campaign_success', campaign_id=campaign_id)


def quick_campaign_builder(request):
    """Hurtig kampagne builder med forudindstillede templates"""
    
    if request.method == 'POST':
        return create_quick_campaign(request)
    
    # Hent industries og template data
    industries = Industry.objects.all()
    template = create_basic_campaign_template()
    
    context = {
        'industries': industries,
        'template': template,
        'quick_mode': True
    }
    
    return render(request, 'campaigns/quick_builder.html', context)


def create_quick_campaign(request):
    """Opret hurtig kampagne baseret på template"""
    try:
        # Basic campaign data
        client_name = request.POST.get('client_name')
        industry_id = request.POST.get('industry')
        website_url = request.POST.get('website_url')
        location = request.POST.get('location', 'Danmark')
        daily_budget = request.POST.get('daily_budget', 500)
        
        # Service keywords
        service_keywords = request.POST.get('service_keywords', '').split(',')
        service_keywords = [k.strip() for k in service_keywords if k.strip()]
        
        # Brand/company name for branded keywords
        brand_name = request.POST.get('brand_name', client_name)
        
        # USPs
        usp1 = request.POST.get('usp1', 'Professionel service')
        usp2 = request.POST.get('usp2', 'Hurtig levering')
        usp3 = request.POST.get('usp3', 'Konkurrencedygtige priser')
        
        # Validation
        if not all([client_name, industry_id, website_url]):
            messages.error(request, 'Udfyld venligst alle påkrævede felter')
            return redirect('quick_campaign')
        
        # Get or create industry
        industry = Industry.objects.get(id=industry_id)
        
        # Create client
        from django.contrib.auth.models import User
        user = request.user if request.user.is_authenticated else User.objects.first()
        
        client, created = Client.objects.get_or_create(
            name=client_name,
            website_url=website_url,
            defaults={
                'industry': industry,
                'description': f'Hurtig kampagne for {industry.name}',
                'created_by': user
            }
        )
        
        # Create campaign
        campaign = Campaign.objects.create(
            name=f"{client_name} - Search Kampagne",
            client=client,
            campaign_type='search',
            budget_daily=daily_budget,
            target_location=location,
            status='draft'
        )
        
        # Create ad groups and content
        create_quick_campaign_structure(campaign, client, service_keywords, brand_name, usp1, usp2, usp3)
        
        messages.success(request, f'Hurtig kampagne "{campaign.name}" er oprettet!')
        return redirect('campaign_success', campaign_id=campaign.id)
        
    except Exception as e:
        messages.error(request, f'Fejl ved oprettelse: {str(e)}')
        return redirect('quick_campaign')


def create_quick_campaign_structure(campaign, client, service_keywords, brand_name, usp1, usp2, usp3):
    """Opret kampagne struktur baseret på hurtig template"""
    
    # Brand Ad Group
    brand_adgroup = AdGroup.objects.create(
        name="Brand Keywords",
        campaign=campaign,
        default_cpc=8.00,
        priority_score=100
    )
    
    # Brand keywords
    brand_keywords = [
        brand_name.lower(),
        f"{brand_name.lower()} {client.industry.name.lower()}",
        f"{brand_name.lower()} {campaign.target_location.lower()}"
    ]
    
    for keyword_text in brand_keywords:
        if keyword_text.strip():
            Keyword.objects.create(
                text=keyword_text,
                ad_group=brand_adgroup,
                match_type='exact',
                max_cpc=8.00
            )
    
    # Service Ad Groups
    if service_keywords:
        service_adgroup = AdGroup.objects.create(
            name="Service Keywords", 
            campaign=campaign,
            default_cpc=15.00,
            priority_score=90
        )
        
        for service in service_keywords[:5]:  # Max 5 services
            # Create exact and phrase match
            Keyword.objects.create(
                text=service,
                ad_group=service_adgroup,
                match_type='phrase',
                max_cpc=15.00
            )
            
            # Location based keyword
            location_keyword = f"{service} {campaign.target_location}"
            Keyword.objects.create(
                text=location_keyword,
                ad_group=service_adgroup,
                match_type='phrase', 
                max_cpc=18.00
            )
    
    # Generic industry keywords
    generic_adgroup = AdGroup.objects.create(
        name="Generiske Keywords",
        campaign=campaign,
        default_cpc=12.00,
        priority_score=70
    )
    
    generic_keywords = [
        f"{client.industry.name.lower()}",
        f"{client.industry.name.lower()} {campaign.target_location.lower()}",
        f"god {client.industry.name.lower()}"
    ]
    
    for keyword_text in generic_keywords:
        if keyword_text.strip():
            Keyword.objects.create(
                text=keyword_text,
                ad_group=generic_adgroup,
                match_type='phrase',
                max_cpc=12.00
            )
    
    # Create ads for each ad group
    ad_groups = [brand_adgroup, service_adgroup, generic_adgroup]
    
    for i, ad_group in enumerate(ad_groups):
        if ad_group.name == "Brand Keywords":
            headline1 = f"{brand_name}"[:30]
            headline2 = f"{usp1}"[:30]
            headline3 = f"{campaign.target_location}"[:30]
        else:
            headline1 = f"{client.industry.name} i {campaign.target_location}"[:30]
            headline2 = f"{usp1}"[:30]  
            headline3 = f"{usp2}"[:30]
        
        description1 = f"{usp3} - Ring i dag for gratis tilbud!"[:90]
        description2 = f"Professionel {client.industry.name.lower()} siden mange år"[:90]
        
        Ad.objects.create(
            ad_group=ad_group,
            headline_1=headline1,
            headline_2=headline2,
            headline_3=headline3,
            description_1=description1,
            description_2=description2,
            final_url=client.website_url,
            display_path_1=client.industry.name.lower().replace(' ', '')[:15],
            display_path_2=campaign.target_location.split(',')[0].lower()[:15]
        )


def geo_campaign_builder(request):
    """Geografisk kampagne builder med kort interface"""
    
    if request.method == 'POST':
        return create_geo_campaign(request)
    
    # Hent data til form
    industries = Industry.objects.all()
    geo_templates = GeoTemplate.objects.filter(is_active=True)
    
    context = {
        'industries': industries,
        'geo_templates': geo_templates,
        'google_maps_api_key': 'AIzaSyBDH6MTS0Hq0ISb0bNQjEAC14321pzM0jw',  # Fra dit kort eksempel
    }
    
    return render(request, 'campaigns/geo_builder.html', context)


def geo_campaign_builder_v2(request):
    """New multi-step geografisk kampagne builder"""
    
    if request.method == 'POST':
        return create_geo_campaign_v2(request)
    
    # Hent data til form
    industries = Industry.objects.all()
    geo_templates = GeoTemplate.objects.filter(is_active=True)
    
    # Get active negative keyword lists for selection
    negative_keyword_lists = NegativeKeywordList.objects.filter(
        is_active=True
    ).select_related('created_by').prefetch_related('negative_keywords')
    
    # Get USP categories and templates for selection
    usp_categories = USPMainCategory.objects.filter(
        is_active=True
    ).order_by('sort_order')
    
    usp_templates = USPTemplate.objects.filter(
        is_active=True,
        main_category__is_active=True
    ).select_related('main_category').prefetch_related('ideal_for_industries').order_by(
        'main_category__sort_order', 'priority_rank'
    )
    
    context = {
        'industries': industries,
        'geo_templates': geo_templates,
        'negative_keyword_lists': negative_keyword_lists,
        'usp_categories': usp_categories,
        'usp_templates': usp_templates,
        'google_maps_api_key': 'AIzaSyBDH6MTS0Hq0ISb0bNQjEAC14321pzM0jw',
    }
    
    return render(request, 'campaigns/geo_builder_v2.html', context)


def create_geo_campaign(request):
    """Opret geo kampagne baseret på kort udvalg"""
    try:
        # Basic data
        client_name = request.POST.get('client_name')
        industry_id = request.POST.get('industry')
        website_url = request.POST.get('website_url')
        service_name = request.POST.get('service_name')
        domain = request.POST.get('domain', '')
        
        # Geo data fra kort
        selected_cities = request.POST.get('selected_cities', '')
        cities = [city.strip() for city in selected_cities.split(',') if city.strip()]
        
        # Template data
        template_id = request.POST.get('template_id')
        meta_title_template = request.POST.get('meta_title_template')
        meta_description_template = request.POST.get('meta_description_template')
        
        # Validering
        is_valid, errors = validate_geo_data(service_name, cities)
        if not is_valid:
            for error in errors:
                messages.error(request, error)
            return redirect('geo_campaign_builder')
        
        if not all([client_name, industry_id, website_url, service_name]):
            messages.error(request, 'Alle påkrævede felter skal udfyldes')
            return redirect('geo_campaign_builder')
        
        # Opret eller hent template
        if template_id:
            template = GeoTemplate.objects.get(id=template_id)
        else:
            # Opret custom template
            template = GeoTemplate.objects.create(
                name=f"{service_name} Custom",
                service_name=service_name,
                meta_title_template=meta_title_template or f"{service_name} {{BYNAVN}} - 5/5 Stjerner på Trustpilot - Ring idag",
                meta_description_template=meta_description_template or f"Skal du bruge en dygtig {service_name.lower()} i {{BYNAVN}}, vi har hjælpet mere end 500 kunder. Kontakt os idag, vi køre dagligt i {{BYNAVN}}",
            )
        
        # Opret client og industry
        industry = Industry.objects.get(id=industry_id)
        user = request.user if request.user.is_authenticated else User.objects.first()
        
        client, created = Client.objects.get_or_create(
            name=client_name,
            website_url=website_url,
            defaults={
                'industry': industry,
                'description': f'Geo kampagne for {service_name}',
                'created_by': user
            }
        )
        
        # Opret geo kampagne
        campaign_name = f"GEO: {service_name} - {len(cities)} byer"
        campaign, geo_keywords = GeoCampaignManager.create_geo_campaign(
            campaign_name=campaign_name,
            service_name=service_name,
            cities=cities,
            template=template,
            client=client,
            user=user
        )
        
        messages.success(request, f'Geo kampagne "{campaign_name}" oprettet med {len(geo_keywords)} keywords!')
        return redirect('geo_campaign_success', campaign_id=campaign.id)
        
    except Exception as e:
        messages.error(request, f'Fejl ved oprettelse af geo kampagne: {str(e)}')
        return redirect('geo_campaign_builder')


def geo_campaign_success(request, campaign_id):
    """Success side for geo kampagner"""
    try:
        campaign = Campaign.objects.get(id=campaign_id)
        geo_keywords = GeoKeyword.objects.filter(campaign=campaign)
        
        if not geo_keywords.exists():
            messages.error(request, 'Ingen geo keywords fundet for kampagnen')
            return redirect('geo_campaign_builder')
        
        # Samlet statistik
        template = geo_keywords.first().template
        cities = [gk.city_name for gk in geo_keywords]
        
        context = {
            'campaign': campaign,
            'template': template,
            'geo_keywords_count': geo_keywords.count(),
            'cities': cities,
            'cities_count': len(cities),
            'service_name': template.service_name,
            'sample_keywords': geo_keywords[:5],
            'sample_urls': [gk.final_url for gk in geo_keywords[:3]],
        }
        
        return render(request, 'campaigns/geo_success.html', context)
        
    except Campaign.DoesNotExist:
        messages.error(request, 'Geo kampagne ikke fundet')
        return redirect('geo_campaign_builder')


@csrf_exempt
def export_geo_campaign(request, campaign_id, export_type):
    """Eksporter geo kampagne i forskellige formater"""
    try:
        campaign = Campaign.objects.get(id=campaign_id)
        
        # Verificer at det er en geo kampagne
        geo_keywords = GeoKeyword.objects.filter(campaign=campaign)
        if not geo_keywords.exists():
            messages.error(request, 'Dette er ikke en geo kampagne')
            return redirect('campaign_success', campaign_id=campaign_id)
        
        # Eksporter
        response = GeoCampaignManager.export_geo_campaign(campaign, export_type)
        
        # Log eksport
        template = geo_keywords.first().template
        cities = [gk.city_name for gk in geo_keywords]
        
        GeoExport.objects.create(
            campaign=campaign,
            template=template,
            export_type=export_type,
            cities_exported=cities,
            keywords_count=geo_keywords.count(),
            exported_by=request.user if request.user.is_authenticated else None
        )
        
        return response
        
    except Campaign.DoesNotExist:
        messages.error(request, 'Kampagne ikke fundet')
        return redirect('geo_campaign_builder')
    except Exception as e:
        messages.error(request, f'Fejl ved eksport: {str(e)}')
        return redirect('geo_campaign_success', campaign_id=campaign_id)


@csrf_exempt 
def preview_geo_data(request):
    """AJAX endpoint til at preview geo data"""
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Kun POST requests tilladt'})
    
    try:
        service_name = request.POST.get('service_name')
        cities_string = request.POST.get('cities')
        domain = request.POST.get('domain', '')
        
        if not service_name or not cities_string:
            return JsonResponse({'error': 'Service navn og byer er påkrævet'})
        
        cities = [city.strip() for city in cities_string.split(',') if city.strip()]
        
        # Generer preview data
        generator = GeoKeywordGenerator(service_name, cities[:5], domain)  # Max 5 for preview
        keywords_data = generator.generate_keywords_data()
        
        # Sample WordPress data
        template = create_demo_geo_template(service_name)
        wp_data = generator.generate_wordpress_data(template)
        
        return JsonResponse({
            'success': True,
            'service_name': service_name,
            'cities_count': len(cities),
            'keywords_sample': keywords_data[:3],
            'wordpress_sample': wp_data[:3],
            'total_keywords': len(cities),
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Preview fejl: {str(e)}'})


def test_geo_system(request):
    """Test side til geo systemet"""
    
    # Test med Fugemand eksempel data
    service_name = "Fugemand"
    cities = ["Bagsværd", "Ølstykke", "Måløv", "Frederikssund", "Furesø"]
    domain = "lundsfugeservice.dk"
    
    # Opret template
    template = create_demo_geo_template(service_name)
    
    # Test generator
    generator = GeoKeywordGenerator(service_name, cities, domain)
    keywords_data = generator.generate_keywords_data()
    wordpress_data = generator.generate_wordpress_data(template)
    
    context = {
        'service_name': service_name,
        'cities': cities,
        'domain': domain,
        'template': template,
        'keywords_data': keywords_data,
        'wordpress_data': wordpress_data,
    }
    
    return render(request, 'campaigns/geo_test.html', context)


def create_geo_campaign_v2(request):
    """Create geo campaign from multi-step form"""
    try:
        # Step 1: Campaign settings
        client_name = request.POST.get('client_name')
        industry_id = request.POST.get('industry') 
        website_url = request.POST.get('website_url')
        service_name = request.POST.get('service_name')
        domain = request.POST.get('domain', '')
        budget_daily = request.POST.get('budget_daily', 500)
        budget_type = request.POST.get('budget_type', 'daily')
        ad_rotation = request.POST.get('ad_rotation', 'optimize')
        bidding_strategy = request.POST.get('bidding_strategy', 'enhanced_cpc')
        default_bid = request.POST.get('default_bid', 15.00)
        target_cpa = request.POST.get('target_cpa')
        target_roas = request.POST.get('target_roas')
        default_match_type = request.POST.get('default_match_type', 'phrase')
        
        # Step 2: Geography
        selected_cities = request.POST.get('selected_cities', '')
        cities = [city.strip() for city in selected_cities.split(',') if city.strip()]
        
        # Step 3: USP Workshop Data
        selected_usps_data = request.POST.get('selected_usps_data', '{}')
        try:
            usps_data = json.loads(selected_usps_data)
        except json.JSONDecodeError:
            usps_data = {}
        
        # Step 4: Headlines and descriptions
        # All headline templates (1-15)
        headline_templates = {}
        for i in range(1, 16):
            field_name = f'headline_{i}_template'
            default_value = ''
            if i == 1:
                default_value = '{SERVICE} {BYNAVN}'
            elif i == 2:
                default_value = '5/5 Stjerner Trustpilot'
            elif i == 3:
                default_value = 'Ring i dag - Gratis tilbud'
            
            headline_templates[field_name] = request.POST.get(field_name, default_value)
        
        # All description templates (1-4)
        description_templates = {}
        for i in range(1, 5):
            field_name = f'description_{i}_template'
            default_value = ''
            if i == 1:
                default_value = 'Professionel {SERVICE} i {BYNAVN} - Ring i dag for gratis tilbud!'
            elif i == 2:
                default_value = 'Erfaren {SERVICE} med 5/5 stjerner. Vi dækker {BYNAVN} og omegn.'
            
            description_templates[field_name] = request.POST.get(field_name, default_value)
        
        # Meta templates
        meta_title_template = request.POST.get('meta_title_template', '{SERVICE} {BYNAVN} - 5/5 Stjerner på Trustpilot - Ring idag')
        meta_description_template = request.POST.get('meta_description_template', 'Skal du bruge en dygtig {SERVICE} i {BYNAVN}, vi har hjælpet mere end 500 kunder. Kontakt os idag, vi køre dagligt i {BYNAVN}')
        
        # Validation
        is_valid, errors = validate_geo_data(service_name, cities)
        if not is_valid:
            for error in errors:
                messages.error(request, error)
            return redirect('geo_campaign_builder_v2')
            
        if not all([client_name, industry_id, website_url, service_name]):
            messages.error(request, 'Alle påkrævede felter skal udfyldes')
            return redirect('geo_campaign_builder_v2')
            
        # Validate headlines and descriptions character limits
        test_data = {'SERVICE': 'Test', 'BYNAVN': 'Test'}
        validation_errors = []
        
        # Validate first 3 headlines (required)
        for i in range(1, 4):
            field_name = f'headline_{i}_template'
            template = headline_templates.get(field_name, '')
            if template:
                try:
                    processed = template.replace('{SERVICE}', test_data['SERVICE']).replace('{BYNAVN}', test_data['BYNAVN'])
                    if len(processed) > 30:
                        validation_errors.append(f'Headline {i}: For lang ({len(processed)} karakterer, max 30)')
                except Exception:
                    validation_errors.append(f'Headline {i}: Ugyldig template')
        
        # Validate first 2 descriptions (required)
        for i in range(1, 3):
            field_name = f'description_{i}_template'
            template = description_templates.get(field_name, '')
            if template:
                try:
                    processed = template.replace('{SERVICE}', test_data['SERVICE']).replace('{BYNAVN}', test_data['BYNAVN'])
                    if len(processed) > 90:
                        validation_errors.append(f'Description {i}: For lang ({len(processed)} karakterer, max 90)')
                except Exception:
                    validation_errors.append(f'Description {i}: Ugyldig template')
        
        if validation_errors:
            for error in validation_errors:
                messages.error(request, error)
            return redirect('geo_campaign_builder_v2')
        
        # Create geo template with enhanced fields
        template = GeoTemplate.objects.create(
            name=f"{service_name} V2 Template",
            service_name=service_name,
            meta_title_template=meta_title_template,
            meta_description_template=meta_description_template,
            # Set all headline templates
            headline_1_template=headline_templates.get('headline_1_template', ''),
            headline_2_template=headline_templates.get('headline_2_template', ''),
            headline_3_template=headline_templates.get('headline_3_template', ''),
            headline_4_template=headline_templates.get('headline_4_template', ''),
            headline_5_template=headline_templates.get('headline_5_template', ''),
            headline_6_template=headline_templates.get('headline_6_template', ''),
            headline_7_template=headline_templates.get('headline_7_template', ''),
            headline_8_template=headline_templates.get('headline_8_template', ''),
            headline_9_template=headline_templates.get('headline_9_template', ''),
            headline_10_template=headline_templates.get('headline_10_template', ''),
            headline_11_template=headline_templates.get('headline_11_template', ''),
            headline_12_template=headline_templates.get('headline_12_template', ''),
            headline_13_template=headline_templates.get('headline_13_template', ''),
            headline_14_template=headline_templates.get('headline_14_template', ''),
            headline_15_template=headline_templates.get('headline_15_template', ''),
            
            # Set all description templates
            description_1_template=description_templates.get('description_1_template', ''),
            description_2_template=description_templates.get('description_2_template', ''),
            description_3_template=description_templates.get('description_3_template', ''),
            description_4_template=description_templates.get('description_4_template', ''),
            
            default_match_type=default_match_type,
        )
        
        # Create client and campaign with enhanced settings
        industry = Industry.objects.get(id=industry_id)
        from django.contrib.auth.models import User
        user = request.user if request.user.is_authenticated else User.objects.first()
        
        client, created = Client.objects.get_or_create(
            name=client_name,
            website_url=website_url,
            defaults={
                'industry': industry,
                'description': f'Multi-step geo kampagne for {service_name}',
                'created_by': user
            }
        )
        
        # Create campaign with enhanced settings
        campaign_name = f"GEO V2: {service_name} - {len(cities)} byer"
        campaign = Campaign.objects.create(
            name=campaign_name,
            client=client,
            campaign_type='search',
            budget_daily=budget_daily,
            budget_type=budget_type,
            target_location='Multi-geo',
            ad_rotation=ad_rotation,
            bidding_strategy=bidding_strategy,
            default_bid=default_bid,
            target_cpa=float(target_cpa) if target_cpa else None,
            target_roas=float(target_roas) if target_roas else None,
            status='draft'
        )
        
        # Process and save USPs from USP Workshop
        usps_created = save_campaign_usps(campaign, client, usps_data)
        
        # Generate geo keywords using enhanced template
        from .geo_export import GeoCampaignManager
        geo_keywords = GeoCampaignManager.create_geo_keywords(
            campaign=campaign,
            template=template,
            cities=cities,
            domain=domain
        )
        
        messages.success(request, f'Multi-step geo kampagne "{campaign_name}" oprettet med {len(geo_keywords)} keywords og {usps_created} USPs!')
        return redirect('geo_campaign_success', campaign_id=campaign.id)
        
    except Exception as e:
        messages.error(request, f'Fejl ved oprettelse af geo kampagne: {str(e)}')
        return redirect('geo_campaign_builder_v2')


def save_campaign_usps(campaign, client, usps_data):
    """
    Process and save USPs from USP Workshop data
    
    Args:
        campaign: Campaign object
        client: Client object  
        usps_data: Dict with USP data from frontend
        
    Returns:
        int: Number of USPs created
    """
    from usps.models import ClientUSP, USPTemplate
    
    usps_created = 0
    
    for key, usp_info in usps_data.items():
        try:
            # Extract USP information
            template_id = usp_info.get('template_id')
            original_text = usp_info.get('original_text') 
            current_text = usp_info.get('current_text', '')
            is_modified = usp_info.get('is_modified', False)
            is_custom = usp_info.get('is_custom', False)
            
            # Skip empty USPs
            if not current_text or not current_text.strip():
                continue
                
            # Get template reference if not custom
            usp_template = None
            if not is_custom and template_id:
                try:
                    usp_template = USPTemplate.objects.get(id=template_id)
                except USPTemplate.DoesNotExist:
                    usp_template = None
            
            # Create ClientUSP
            client_usp = ClientUSP.objects.create(
                client=client,
                campaign=campaign,
                usp_template=usp_template,
                custom_text=current_text.strip(),
                original_text=original_text,
                is_modified=is_modified,
                is_custom=is_custom,
                is_selected=True
            )
            
            usps_created += 1
            
        except Exception as e:
            print(f"Error saving USP {key}: {e}")
            continue
    
    return usps_created


# =========================
# NEGATIVE KEYWORDS SYSTEM
# =========================

@login_required
def negative_keywords_dashboard(request):
    """Dashboard til negative keywords administration"""
    keyword_lists = NegativeKeywordList.objects.filter(
        created_by=request.user
    ).prefetch_related('negative_keywords')
    
    context = {
        'keyword_lists': keyword_lists,
        'total_keywords': sum(kl.keywords_count for kl in keyword_lists),
        'active_lists': keyword_lists.filter(is_active=True).count(),
    }
    
    return render(request, 'campaigns/negative_keywords_dashboard.html', context)


@login_required
def create_negative_keyword_list(request):
    """Opret ny negative keyword liste"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            category = request.POST.get('category', 'general')
            description = request.POST.get('description', '')
            is_active = request.POST.get('is_active') == 'on'
            auto_apply_industries = request.POST.getlist('auto_apply_industries')
            
            keyword_list = NegativeKeywordList.objects.create(
                name=name,
                category=category,
                description=description,
                is_active=is_active,
                auto_apply_to_industries=auto_apply_industries,
                created_by=request.user
            )
            
            messages.success(request, f'Negative keyword liste "{name}" er oprettet!')
            return redirect('negative_keyword_list_detail', list_id=keyword_list.id)
            
        except Exception as e:
            messages.error(request, f'Fejl ved oprettelse: {str(e)}')
    
    # Get available industries for auto-apply
    industries = Industry.objects.all().values_list('name', flat=True)
    
    context = {
        'industries': industries,
        'category_choices': NegativeKeywordList.CATEGORY_CHOICES,
    }
    
    return render(request, 'campaigns/create_negative_keyword_list.html', context)


@login_required
def negative_keyword_list_detail(request, list_id):
    """Detaljer for en specifik negative keyword liste"""
    keyword_list = get_object_or_404(
        NegativeKeywordList, 
        id=list_id, 
        created_by=request.user
    )
    
    # Get keywords with pagination
    from django.core.paginator import Paginator
    keywords = keyword_list.negative_keywords.all()
    paginator = Paginator(keywords, 50)  # 50 keywords per side
    page = request.GET.get('page', 1)
    keywords_page = paginator.get_page(page)
    
    # Get upload history
    uploads = NegativeKeywordUpload.objects.filter(
        keyword_list=keyword_list
    ).order_by('-uploaded_at')[:10]
    
    context = {
        'keyword_list': keyword_list,
        'keywords': keywords_page,
        'uploads': uploads,
        'match_type_choices': NegativeKeyword.MATCH_TYPES,
    }
    
    return render(request, 'campaigns/negative_keyword_list_detail.html', context)


@login_required
def upload_negative_keywords(request, list_id):
    """Upload negative keywords fra fil"""
    keyword_list = get_object_or_404(
        NegativeKeywordList, 
        id=list_id, 
        created_by=request.user
    )
    
    if request.method == 'POST' and request.FILES.get('keyword_file'):
        try:
            uploaded_file = request.FILES['keyword_file']
            
            # Validér fil type
            if not uploaded_file.name.lower().endswith(('.txt', '.csv')):
                messages.error(request, 'Kun .txt og .csv filer er tilladt')
                return redirect('negative_keyword_list_detail', list_id=list_id)
            
            # Process file upload
            result = process_negative_keyword_file(
                uploaded_file, 
                keyword_list, 
                request.user
            )
            
            if result['success']:
                messages.success(
                    request, 
                    f'Upload færdig! {result["keywords_added"]} keywords tilføjet, {result["keywords_skipped"]} sprunget over'
                )
            else:
                messages.error(request, f'Upload fejl: {result["error"]}')
                
        except Exception as e:
            messages.error(request, f'Fejl ved upload: {str(e)}')
    
    return redirect('negative_keyword_list_detail', list_id=list_id)


def process_negative_keyword_file(uploaded_file, keyword_list, user):
    """Process uploaded negative keyword fil"""
    import csv
    import io
    from django.utils import timezone
    
    try:
        # Create upload record
        upload_record = NegativeKeywordUpload.objects.create(
            keyword_list=keyword_list,
            original_filename=uploaded_file.name,
            file_size_kb=uploaded_file.size // 1024,
            uploaded_by=user,
            status='processing'
        )
        
        # Read file content
        file_content = uploaded_file.read().decode('utf-8')
        lines = file_content.strip().split('\n')
        
        keywords_added = 0
        keywords_skipped = 0
        keywords_errors = 0
        error_details = []
        
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if not line or line.startswith('#'):
                continue
                
            try:
                # Parse keyword and match type
                keyword_text, match_type = parse_negative_keyword_line(line)
                
                # Check if keyword already exists
                if NegativeKeyword.objects.filter(
                    keyword_list=keyword_list,
                    keyword_text=keyword_text,
                    match_type=match_type
                ).exists():
                    keywords_skipped += 1
                    continue
                
                # Create negative keyword
                NegativeKeyword.objects.create(
                    keyword_list=keyword_list,
                    keyword_text=keyword_text,
                    match_type=match_type,
                    source_file_line=line_num
                )
                keywords_added += 1
                
            except Exception as e:
                keywords_errors += 1
                error_details.append(f'Linje {line_num}: {str(e)}')
        
        # Update upload record
        upload_record.status = 'completed'
        upload_record.completed_at = timezone.now()
        upload_record.total_lines = len(lines)
        upload_record.keywords_added = keywords_added
        upload_record.keywords_skipped = keywords_skipped
        upload_record.keywords_errors = keywords_errors
        upload_record.error_details = '\n'.join(error_details)
        upload_record.save()
        
        # Update keyword list count
        keyword_list.update_keywords_count()
        
        return {
            'success': True,
            'keywords_added': keywords_added,
            'keywords_skipped': keywords_skipped,
            'keywords_errors': keywords_errors
        }
        
    except Exception as e:
        # Update upload record with error
        upload_record.status = 'failed'
        upload_record.error_details = str(e)
        upload_record.completed_at = timezone.now()
        upload_record.save()
        
        return {
            'success': False,
            'error': str(e)
        }


def parse_negative_keyword_line(line):
    """Parse en linje fra negative keyword fil"""
    line = line.strip()
    
    # Remove leading minus if present
    if line.startswith('-'):
        line = line[1:].strip()
    
    # Determine match type based on symbols
    if line.startswith('[') and line.endswith(']'):
        match_type = 'exact'
        keyword_text = line[1:-1].strip()
    elif line.startswith('"') and line.endswith('"'):
        match_type = 'phrase'
        keyword_text = line[1:-1].strip()
    else:
        match_type = 'broad'
        keyword_text = line.strip()
    
    if not keyword_text:
        raise ValueError("Tomt keyword")
    
    return keyword_text, match_type


@login_required
def get_campaign_negative_keywords(request, campaign_id):
    """AJAX view til at hente negative keywords for kampagne"""
    campaign = get_object_or_404(Campaign, id=campaign_id)
    
    # Get active negative keyword lists for this campaign
    negative_lists = NegativeKeywordList.objects.filter(
        campaignNegativekeywordlist__campaign=campaign,
        campaignNegativekeywordlist__is_active=True
    ).prefetch_related('negative_keywords')
    
    # Also get auto-applied lists based on industry
    auto_lists = NegativeKeywordList.objects.filter(
        is_active=True,
        auto_apply_to_industries__contains=[campaign.client.industry.name]
    ).prefetch_related('negative_keywords')
    
    # Combine all keywords
    all_keywords = []
    
    for neg_list in list(negative_lists) + list(auto_lists):
        for keyword in neg_list.negative_keywords.all():
            all_keywords.append({
                'text': keyword.keyword_text,
                'match_type': keyword.match_type,
                'list_name': neg_list.name,
                'formatted': str(keyword)
            })
    
    return JsonResponse({
        'keywords': all_keywords,
        'total_count': len(all_keywords)
    })


@login_required  
def apply_negative_keywords_to_campaign(request, campaign_id):
    """Anvend negative keyword lister til kampagne"""
    if request.method == 'POST':
        campaign = get_object_or_404(Campaign, id=campaign_id)
        selected_lists = request.POST.getlist('negative_lists')
        
        try:
            # Remove existing associations
            CampaignNegativeKeywordList.objects.filter(campaign=campaign).delete()
            
            # Add new associations
            for list_id in selected_lists:
                negative_list = NegativeKeywordList.objects.get(id=list_id)
                CampaignNegativeKeywordList.objects.create(
                    campaign=campaign,
                    negative_list=negative_list,
                    applied_by=request.user
                )
            
            messages.success(
                request, 
                f'{len(selected_lists)} negative keyword lister tilføjet til kampagne'
            )
            
        except Exception as e:
            messages.error(request, f'Fejl ved tilføjelse: {str(e)}')
    
    return redirect('campaign_detail', campaign_id=campaign_id)


# ====================================
# MODERN NEGATIVE KEYWORDS MANAGER - AJAX ENDPOINTS
# ====================================

@csrf_exempt
def negative_keywords_manager(request):
    """Modern negative keywords manager with enhanced UI"""
    if request.user.is_authenticated:
        keyword_lists = NegativeKeywordList.objects.filter(
            created_by=request.user
        ).prefetch_related('negative_keywords').select_related('industry').order_by('-created_at')
    else:
        # For demo purposes, show all lists when not authenticated
        keyword_lists = NegativeKeywordList.objects.all().prefetch_related('negative_keywords').select_related('industry').order_by('-created_at')
    
    # Get all industries for filter dropdown
    industries = Industry.objects.all().order_by('name')
    
    # Calculate statistics
    categories_used = set(kl.category for kl in keyword_lists)
    industries_used = set(kl.industry for kl in keyword_lists if kl.industry)
    
    context = {
        'keyword_lists': keyword_lists,
        'industries': industries,
        'total_keywords': sum(kl.keywords_count for kl in keyword_lists),
        'active_lists': keyword_lists.filter(is_active=True).count(),
        'categories_count': len(categories_used),
        'industries_count': len(industries_used),
    }
    
    return render(request, 'campaigns/negative_keywords_manager.html', context)


@csrf_exempt
def create_negative_keyword_list_ajax(request):
    """AJAX endpoint to create new negative keyword list"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            category = request.POST.get('category', 'general')
            description = request.POST.get('description', '').strip()
            is_active = request.POST.get('is_active') == 'true'
            industry_id = request.POST.get('industry', '').strip()
            icon = request.POST.get('icon', '📋').strip()
            color = request.POST.get('color', '#8B5CF6').strip()
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Liste navn er påkrævet'})
            
            # Get user for creation - use authenticated user or demo user as fallback  
            if request.user.is_authenticated:
                created_by = request.user
            else:
                # Fallback to demo user for testing
                from django.contrib.auth.models import User
                created_by = User.objects.get(username='demo')
            
            # Check if name already exists for this user
            if NegativeKeywordList.objects.filter(
                name__iexact=name, 
                created_by=created_by
            ).exists():
                return JsonResponse({'success': False, 'error': 'En liste med dette navn eksisterer allerede'})
            
            # Get industry if provided
            industry = None
            if industry_id and industry_id.strip() and industry_id.strip() != '':
                try:
                    industry_id_int = int(industry_id.strip())
                    industry = Industry.objects.get(id=industry_id_int)
                except (ValueError, Industry.DoesNotExist):
                    return JsonResponse({'success': False, 'error': 'Ugyldig branche valgt'})
            
            keyword_list = NegativeKeywordList.objects.create(
                name=name,
                category=category,
                description=description,
                industry=industry,
                icon=icon,
                color=color,
                is_active=is_active or True,  # Default to active
                created_by=created_by,
                auto_apply_to_industries=[]
            )
            
            # Handle initial keywords if provided
            from .models import NegativeKeyword
            import json
            initial_keywords = request.POST.get('initial_keywords', '').strip()
            keywords_added = 0

            if initial_keywords:
                # Try to parse as JSON first (from industry_manager)
                try:
                    keywords_data = json.loads(initial_keywords)
                    if isinstance(keywords_data, list):
                        for kw in keywords_data:
                            keyword_text = kw.get('keyword_text', '').strip().lower() if isinstance(kw, dict) else str(kw).strip().lower()
                            match_type = kw.get('match_type', 'broad') if isinstance(kw, dict) else 'broad'

                            if keyword_text:  # Only create if keyword text is not empty
                                if not keyword_list.negative_keywords.filter(keyword_text__iexact=keyword_text).exists():
                                    NegativeKeyword.objects.create(
                                        keyword_list=keyword_list,
                                        keyword_text=keyword_text,
                                        match_type=match_type
                                    )
                                    keywords_added += 1
                except json.JSONDecodeError:
                    # Fallback: Split by newlines and commas (plain text input)
                    import re
                    keywords_list = re.split(r'[,\n]+', initial_keywords)
                    keywords_list = [k.strip().lower() for k in keywords_list if k.strip()]

                    for keyword_text in keywords_list:
                        if keyword_text:  # Only create if keyword text is not empty
                            if not keyword_list.negative_keywords.filter(keyword_text__iexact=keyword_text).exists():
                                NegativeKeyword.objects.create(
                                    keyword_list=keyword_list,
                                    keyword_text=keyword_text,
                                    match_type='broad'
                                )
                                keywords_added += 1
            
            return JsonResponse({
                'success': True,
                'list_id': keyword_list.id,  # Return list_id as expected by JS
                'message': f'Liste "{name}" oprettet succesfuldt' + (f' med {keywords_added} søgeord' if keywords_added > 0 else ''),
                'list': {
                    'id': keyword_list.id,
                    'name': keyword_list.name,
                    'category': keyword_list.category,
                    'description': keyword_list.description,
                    'is_active': keyword_list.is_active,
                    'keywords_count': keywords_added
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@csrf_exempt  
def add_negative_keyword_ajax(request):
    """AJAX endpoint to add new keyword to list"""
    if request.method == 'POST':
        try:
            list_id = request.POST.get('list_id')
            keyword_text = request.POST.get('keyword_text', '').strip()
            match_type = request.POST.get('match_type', 'broad')
            
            if not keyword_text:
                return JsonResponse({'success': False, 'error': 'Søgeord er påkrævet'})
            
            # Get the list (allow access for demo mode)
            if request.user.is_authenticated:
                keyword_list = get_object_or_404(
                    NegativeKeywordList, 
                    id=list_id, 
                    created_by=request.user
                )
            else:
                # Demo mode - allow access to any list
                keyword_list = get_object_or_404(NegativeKeywordList, id=list_id)
            
            # Intelligent hierarkisk analyse som ved Excel import
            from .services import NegativeKeywordConflictAnalyzer
            analyzer = NegativeKeywordConflictAnalyzer(keyword_list)
            
            # Normaliser keyword til analyzer format
            normalized_kw = {
                'text': keyword_text.lower().strip(),
                'original_text': keyword_text,
                'match_type': match_type
            }
            
            # Analyser relationerne
            relationships = analyzer._analyze_all_relationships(normalized_kw)
            removed_keywords = []
            
            if relationships['identical']:
                # Identisk keyword eksisterer allerede
                existing = relationships['identical'][0]
                return JsonResponse({
                    'success': False, 
                    'error': f'"{existing["original_text"]}" ({existing["match_type"]}) eksisterer allerede i listen'
                })
                
            elif relationships['blocked_by']:
                # Blokeret af højere hierarki
                blocking_kw = relationships['blocked_by'][0]
                return JsonResponse({
                    'success': False, 
                    'error': f'Blokeret af eksisterende "{blocking_kw["original_text"]}" ({blocking_kw["match_type"]}) - denne har højere hierarki og dækker allerede dette keyword'
                })
            
            else:
                # Safe to add - men check for cleanup opportunities
                if relationships['will_override']:
                    # Fjern redundante keywords først
                    for override_kw in relationships['will_override']:
                        existing_to_remove = NegativeKeyword.objects.filter(
                            id=override_kw['id'],
                            keyword_list=keyword_list
                        ).first()
                        
                        if existing_to_remove:
                            removed_keywords.append({
                                'text': existing_to_remove.keyword_text,
                                'match_type': existing_to_remove.match_type
                            })
                            existing_to_remove.delete()
                
                # Create the keyword
                keyword = NegativeKeyword.objects.create(
                    keyword_list=keyword_list,
                    keyword_text=keyword_text,
                    match_type=match_type
                )
                
                # Opdater keyword count
                keyword_list.update_keywords_count()
                
                response_data = {
                    'success': True,
                    'keyword': {
                        'id': keyword.id,
                        'text': keyword.keyword_text,
                        'match_type': keyword.match_type,
                        'match_type_display': keyword.get_match_type_display(),
                        'added_at': keyword.added_at.strftime('%d/%m %Y')
                    }
                }
                
                # Tilføj cleanup information hvis relevant
                if removed_keywords:
                    response_data['removed_keywords'] = removed_keywords
                    response_data['message'] = f'Tilføjet "{keyword_text}" ({match_type}) og fjernet {len(removed_keywords)} redundante keywords'
                
                return JsonResponse(response_data)
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@csrf_exempt
def delete_negative_keyword_ajax(request, keyword_id):
    """AJAX endpoint to delete a negative keyword"""
    if request.method == 'POST':
        try:
            if request.user.is_authenticated:
                keyword = get_object_or_404(
                    NegativeKeyword, 
                    id=keyword_id,
                    keyword_list__created_by=request.user
                )
            else:
                # Demo mode - allow access to any keyword
                keyword = get_object_or_404(NegativeKeyword, id=keyword_id)
            
            keyword_text = keyword.keyword_text
            keyword.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Søgeordet "{keyword_text}" er slettet'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@csrf_exempt
def update_negative_keyword_ajax(request, keyword_id):
    """AJAX endpoint to update a negative keyword"""
    if request.method == 'POST':
        try:
            if request.user.is_authenticated:
                keyword = get_object_or_404(
                    NegativeKeyword, 
                    id=keyword_id,
                    keyword_list__created_by=request.user
                )
            else:
                # Demo mode - allow access to any keyword
                keyword = get_object_or_404(NegativeKeyword, id=keyword_id)
            
            # Get new values
            new_keyword_text = request.POST.get('keyword_text', '').strip()
            new_match_type = request.POST.get('match_type', 'broad')
            
            # Validate inputs
            if not new_keyword_text:
                return JsonResponse({
                    'success': False, 
                    'error': 'Søgeord kan ikke være tomt'
                })
            
            if new_match_type not in ['broad', 'phrase', 'exact']:
                return JsonResponse({
                    'success': False, 
                    'error': 'Ugyldig match type'
                })
            
            # Check for duplicates in the same list (excluding current keyword)
            existing_keyword = NegativeKeyword.objects.filter(
                keyword_list=keyword.keyword_list,
                keyword_text__iexact=new_keyword_text,
                match_type=new_match_type
            ).exclude(id=keyword_id).first()
            
            if existing_keyword:
                return JsonResponse({
                    'success': False,
                    'error': f'Søgeordet "{new_keyword_text}" med {new_match_type} match findes allerede i listen'
                })
            
            # Update the keyword
            old_text = keyword.keyword_text
            old_match = keyword.match_type
            
            keyword.keyword_text = new_keyword_text
            keyword.match_type = new_match_type
            keyword.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Søgeord opdateret fra "{old_text}" ({old_match}) til "{new_keyword_text}" ({new_match_type})',
                'keyword': {
                    'id': keyword.id,
                    'text': keyword.keyword_text,
                    'match_type': keyword.match_type,
                    'match_type_display': keyword.get_match_type_display()
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@csrf_exempt
def delete_negative_keyword_list_ajax(request, list_id):
    """AJAX endpoint to delete an entire negative keyword list"""
    if request.method == 'POST':
        try:
            # Get the list by ID only (no user filter for demo/anonymous users)
            keyword_list = get_object_or_404(
                NegativeKeywordList,
                id=list_id
            )
            
            # Count keywords before deletion
            keywords_count = keyword_list.negative_keywords.count()
            list_name = keyword_list.name
            
            # Delete the list (keywords will be deleted automatically due to CASCADE)
            keyword_list.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Listen "{list_name}" og {keywords_count} søgeord er slettet'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@csrf_exempt
def edit_negative_keyword_list_ajax(request, list_id):
    """AJAX endpoint to edit negative keyword list"""
    if request.method == 'POST':
        try:
            # Get the list by ID only (no user filter for demo/anonymous users)
            keyword_list = get_object_or_404(
                NegativeKeywordList,
                id=list_id
            )
            
            # Update fields
            keyword_list.name = request.POST.get('name', keyword_list.name).strip()
            keyword_list.category = request.POST.get('category', keyword_list.category)
            keyword_list.description = request.POST.get('description', keyword_list.description).strip()
            keyword_list.is_active = request.POST.get('is_active', 'false') == 'true'
            keyword_list.icon = request.POST.get('icon', keyword_list.icon).strip()
            keyword_list.color = request.POST.get('color', keyword_list.color).strip()
            
            # Update industry if provided
            industry_id = request.POST.get('industry')
            if industry_id:
                try:
                    from .models import Industry
                    industry = Industry.objects.get(id=industry_id)
                    keyword_list.industry = industry
                except Industry.DoesNotExist:
                    keyword_list.industry = None
            else:
                keyword_list.industry = None
            
            keyword_list.save()
            
            # Force refresh from database to verify save
            keyword_list.refresh_from_db()
            
            return JsonResponse({
                'success': True,
                'list': {
                    'id': keyword_list.id,
                    'name': keyword_list.name,
                    'category': keyword_list.category,
                    'description': keyword_list.description,
                    'is_active': keyword_list.is_active
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@csrf_exempt 
def get_negative_keyword_list_ajax(request, list_id):
    """AJAX endpoint to get negative keyword list data"""
    if request.method == 'GET':
        try:
            # Get the list by ID only (no user filter for demo/anonymous users)
            keyword_list = get_object_or_404(
                NegativeKeywordList,
                id=list_id
            )
            
            return JsonResponse({
                'success': True,
                'list': {
                    'id': keyword_list.id,
                    'name': keyword_list.name,
                    'category': keyword_list.category,
                    'description': keyword_list.description,
                    'is_active': keyword_list.is_active,
                    'keywords_count': keyword_list.keywords_count
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@csrf_exempt
def get_negative_keywords_for_list_ajax(request, list_id):
    """AJAX endpoint to get all keywords for a negative keyword list"""
    if request.method == 'GET':
        try:
            # Get the list (allow access for demo mode)
            if request.user.is_authenticated:
                keyword_list = get_object_or_404(
                    NegativeKeywordList,
                    id=list_id,
                    created_by=request.user
                )
            else:
                # Demo mode - allow access to any list
                keyword_list = get_object_or_404(NegativeKeywordList, id=list_id)

            # Get all keywords for this list
            keywords = keyword_list.negative_keywords.all().order_by('-added_at')

            keywords_data = []
            for kw in keywords:
                match_type_display = {
                    'broad': 'Broad Match',
                    'phrase': 'Phrase Match',
                    'exact': 'Exact Match'
                }.get(kw.match_type, kw.match_type.title())

                keywords_data.append({
                    'id': kw.id,
                    'keyword_text': kw.keyword_text,
                    'match_type': kw.match_type,
                    'match_type_display': match_type_display,
                    'added_at': kw.added_at.strftime('%d/%m %Y') if kw.added_at else ''
                })

            return JsonResponse({
                'success': True,
                'keywords': keywords_data,
                'list_name': keyword_list.name,
                'keywords_count': len(keywords_data)
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid method'})


@csrf_exempt
def analyze_excel_import_for_list(request, list_id):
    """AJAX endpoint to analyze Excel import for specific list with conflict detection"""
    if request.method == 'POST':
        try:
            from .services import NegativeKeywordConflictAnalyzer
            import openpyxl
            import io
            
            # Get the list (allow access for demo mode)
            if request.user.is_authenticated:
                keyword_list = get_object_or_404(
                    NegativeKeywordList, 
                    id=list_id, 
                    created_by=request.user
                )
            else:
                # Demo mode - allow access to any list
                keyword_list = get_object_or_404(NegativeKeywordList, id=list_id)
            
            # Get uploaded file
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                return JsonResponse({'success': False, 'error': 'Ingen fil uploadet'})
            
            # Parse Excel file
            try:
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active
                
                # Read keywords from Excel
                import_keywords = []
                
                # Skip header row and read data
                for row_num in range(2, worksheet.max_row + 1):
                    keyword_text = worksheet.cell(row=row_num, column=1).value
                    match_type = worksheet.cell(row=row_num, column=2).value
                    
                    if keyword_text and match_type:
                        # Clean and validate data
                        keyword_text = str(keyword_text).strip()
                        match_type = str(match_type).lower().strip()
                        
                        # Normalize match type
                        if match_type in ['broad', 'broad match']:
                            match_type = 'broad'
                        elif match_type in ['phrase', 'phrase match']:
                            match_type = 'phrase'
                        elif match_type in ['exact', 'exact match']:
                            match_type = 'exact'
                        else:
                            continue  # Skip invalid match types
                        
                        if keyword_text:  # Only add non-empty keywords
                            import_keywords.append({
                                'text': keyword_text,
                                'match_type': match_type
                            })
                
                if not import_keywords:
                    return JsonResponse({
                        'success': False, 
                        'error': 'Ingen gyldige keywords fundet i Excel filen. Tjek formatet.'
                    })
                
                # Initialize conflict analyzer
                analyzer = NegativeKeywordConflictAnalyzer(keyword_list)
                
                # Analyze conflicts
                analysis_result = analyzer.analyze_import(import_keywords)
                
                return JsonResponse({
                    'success': True,
                    'analysis': analysis_result,
                    'list_info': {
                        'id': keyword_list.id,
                        'name': keyword_list.name,
                        'existing_count': keyword_list.negative_keywords.count()
                    }
                })
                
            except Exception as e:
                return JsonResponse({
                    'success': False, 
                    'error': f'Fejl ved læsning af Excel fil: {str(e)}'
                })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@csrf_exempt
def cleanup_keywords(request, list_id):
    """AJAX endpoint til at slette specifikke keywords fra listen"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            keyword_ids = data.get('keyword_ids', [])
            
            # Find keyword listen
            keyword_list = get_object_or_404(NegativeKeywordList, id=list_id)
            
            # Udfør cleanup
            from .services import NegativeKeywordConflictAnalyzer
            analyzer = NegativeKeywordConflictAnalyzer(keyword_list)
            result = analyzer.execute_cleanup(keyword_ids)
            
            if result['success']:
                return JsonResponse({
                    'success': True,
                    'removed_count': result['removed_count'],
                    'removed_keywords': result['removed_keywords'],
                    'message': f"Successfuldt slettet {result['removed_count']} keywords"
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': result['error']
                }, status=400)
                
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Ugyldig JSON data'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Fejl under sletning: {str(e)}'
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Kun POST requests tilladt'}, status=405)


@csrf_exempt
def execute_excel_import(request, list_id):
    """AJAX endpoint til at udføre den faktiske import efter konflikt-løsning"""
    if request.method == 'POST':
        try:
            # Check if file was uploaded
            if 'excel_file' not in request.FILES:
                return JsonResponse({
                    'success': False,
                    'error': 'Ingen Excel fil uploadet'
                }, status=400)
            
            excel_file = request.FILES['excel_file']
            keywords_to_add = json.loads(request.POST.get('keywords_to_add', '[]'))
            
            # Find keyword listen
            keyword_list = get_object_or_404(NegativeKeywordList, id=list_id)
            
            # Parse Excel filen igen for at få rå data
            if excel_file.name.endswith('.xlsx'):
                import openpyxl
                workbook = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = workbook.active
                excel_keywords = []
                
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                    if row[0]:  # Hvis der er tekst i første kolonne
                        keyword_text = str(row[0]).strip()
                        match_type_raw = str(row[1]).strip().lower() if row[1] else 'broad'
                        
                        # Normalize match type from Excel format to our format
                        if match_type_raw in ['broad', 'broad match']:
                            match_type = 'broad'
                        elif match_type_raw in ['phrase', 'phrase match']:
                            match_type = 'phrase'
                        elif match_type_raw in ['exact', 'exact match']:
                            match_type = 'exact'
                        else:
                            match_type = 'broad'
                        
                        excel_keywords.append({
                            'text': keyword_text,
                            'match_type': match_type
                        })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Kun .xlsx filer understøttes'
                }, status=400)
            
            # Filtrer keywords baseret på brugerens valg (keywords_to_add)
            keywords_to_import = []
            print(f"[DEBUG] Filtering {len(excel_keywords)} excel keywords against {len(keywords_to_add)} selected")
            print(f"[DEBUG] Selected keywords: {keywords_to_add}")
            
            for kw in excel_keywords:
                # Check if this keyword should be added based on user selection
                # Normalize text to lowercase to match frontend analysis format
                keyword_key = f"{kw['text'].lower().strip()}_{kw['match_type']}"
                print(f"[DEBUG] Checking: {keyword_key}")
                if keyword_key in keywords_to_add:
                    keywords_to_import.append(kw)
                    print(f"[DEBUG] ✅ MATCHED: {keyword_key}")
                else:
                    print(f"[DEBUG] ❌ NOT SELECTED: {keyword_key}")
            
            keyword_list_debug = [f"{kw['text'].lower().strip()}_{kw['match_type']}" for kw in keywords_to_import]
            print(f"[DEBUG] Final keywords to import: {len(keywords_to_import)} - {keyword_list_debug}")
            
            # Intelligent hierarkisk import med auto-cleanup
            added_count = 0
            skipped_count = 0
            removed_count = 0
            errors = []
            removed_keywords = []
            
            # Brug conflict analyzer til hierarkisk analyse
            from .services import NegativeKeywordConflictAnalyzer
            analyzer = NegativeKeywordConflictAnalyzer(keyword_list)
            
            for kw in keywords_to_import:
                try:
                    # Normaliser keyword til format forventet af analyzer
                    normalized_kw = {
                        'text': kw['text'].lower().strip(),
                        'original_text': kw['text'],
                        'match_type': kw['match_type']
                    }
                    
                    # Analyser relationerne
                    relationships = analyzer._analyze_all_relationships(normalized_kw)
                    
                    if relationships['identical']:
                        # Identisk keyword eksisterer allerede - skip
                        skipped_count += 1
                        print(f"[DEBUG] SKIPPED - Identisk: {kw['text']} ({kw['match_type']})")
                        
                    elif relationships['blocked_by']:
                        # Blokeret af højere hierarki - skip
                        blocking_kw = relationships['blocked_by'][0]
                        skipped_count += 1
                        print(f"[DEBUG] BLOCKED - {kw['text']} ({kw['match_type']}) blokeret af {blocking_kw['original_text']} ({blocking_kw['match_type']})")
                        
                    else:
                        # Safe to add eller vil overskrive eksisterende
                        if relationships['will_override']:
                            # Fjern redundante keywords først
                            for override_kw in relationships['will_override']:
                                existing_to_remove = NegativeKeyword.objects.filter(
                                    id=override_kw['id'],
                                    keyword_list=keyword_list
                                ).first()
                                
                                if existing_to_remove:
                                    removed_keywords.append(f"{existing_to_remove.keyword_text} ({existing_to_remove.match_type})")
                                    existing_to_remove.delete()
                                    removed_count += 1
                                    print(f"[DEBUG] REMOVED - {existing_to_remove.keyword_text} ({existing_to_remove.match_type}) - overskrevet af {kw['text']} ({kw['match_type']})")
                        
                        # Tilføj det nye keyword
                        new_keyword = NegativeKeyword.objects.create(
                            keyword_list=keyword_list,
                            keyword_text=kw['text'],
                            match_type=kw['match_type']
                        )
                        added_count += 1
                        print(f"[DEBUG] ADDED - {kw['text']} ({kw['match_type']})")
                        
                        # Opdater analyzer's existing keywords efter tilføjelse
                        analyzer.existing_keywords = analyzer._get_existing_keywords()
                        
                except Exception as e:
                    errors.append(f"Fejl ved tilføjelse af '{kw['text']}': {str(e)}")
                    print(f"[DEBUG] ERROR - {kw['text']}: {str(e)}")
            
            # Opdater keywords count på listen
            keyword_list.update_keywords_count()
            
            print(f"[FINAL DEBUG] Import result: {added_count} added, {skipped_count} skipped, {len(errors)} errors")
            
            # Opret intelligent besked
            message_parts = [f"{added_count} tilføjet"]
            if removed_count > 0:
                message_parts.append(f"{removed_count} fjernet (redundante)")
            if skipped_count > 0:
                message_parts.append(f"{skipped_count} sprunget over")
            
            return JsonResponse({
                'success': True,
                'added_count': added_count,
                'skipped_count': skipped_count,
                'removed_count': removed_count,
                'removed_keywords': removed_keywords,
                'errors': errors,
                'total_keywords_in_list': keyword_list.keywords_count,
                'message': f"Import færdig: {', '.join(message_parts)}"
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Ugyldig JSON data'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Fejl under import: {str(e)}'
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Kun POST requests tilladt'}, status=405)


def download_negative_keywords_template(request):
    """Download Excel template for negative keywords import"""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill
    import io
    
    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Negative Keywords Template"
    
    # Headers - kun 2 kolonner: Søgeord og Match Type
    headers = ['Søgeord', 'Match Type']
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    # Example data - kun søgeord og match type
    examples = [
        ['gratis', 'broad'],
        ['job', 'phrase'],
        ['diy', 'broad'],
        ['billigst', 'broad'],
        ['københavn', 'exact'],
    ]
    
    for row, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            worksheet.cell(row=row, column=col).value = value
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="negative_keywords_template.xlsx"'
    
    # Save workbook to response
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response.write(virtual_workbook.read())
    
    return response


@csrf_exempt
def import_negative_keywords_excel(request):
    """AJAX endpoint to import negative keywords from Excel file"""
    if request.method == 'POST':
        try:
            if 'excel_file' not in request.FILES:
                return JsonResponse({'success': False, 'error': 'Ingen fil uploaded'})
            
            excel_file = request.FILES['excel_file']
            
            # Validate file type
            if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
                return JsonResponse({'success': False, 'error': 'Kun Excel filer (.xlsx, .xls) er understøttet'})
            
            # Process the Excel file
            result = process_negative_keywords_excel(excel_file, request.user)
            
            return JsonResponse(result)
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Fejl ved import: {str(e)}'})
    
    return JsonResponse({'success': False, 'error': 'Invalid method'})


def process_negative_keywords_excel(excel_file, user):
    """Process uploaded Excel file with negative keywords"""
    import openpyxl
    import io
    from collections import defaultdict
    
    try:
        # Read Excel file
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active
        
        # Expected headers
        expected_headers = ['Liste Navn', 'Kategori', 'Søgeord', 'Match Type']
        
        # Get headers from first row
        headers = []
        for col in range(1, 6):  # Check first 5 columns
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
        
        # Check if we have minimum required headers
        missing_headers = []
        for expected in expected_headers:
            if expected not in headers:
                missing_headers.append(expected)
        
        if missing_headers:
            return {
                'success': False, 
                'error': f'Manglende kolonner: {", ".join(missing_headers)}'
            }
        
        # Map header positions
        header_map = {}
        for i, header in enumerate(headers):
            header_map[header] = i + 1
        
        # Process data rows
        lists_data = defaultdict(lambda: {
            'category': 'general',
            'keywords': []
        })
        
        processed_rows = 0
        skipped_rows = 0
        errors = []
        
        for row_num in range(2, worksheet.max_row + 1):
            try:
                # Get cell values
                list_name = worksheet.cell(row=row_num, column=header_map['Liste Navn']).value
                category = worksheet.cell(row=row_num, column=header_map.get('Kategori', 1)).value
                keyword = worksheet.cell(row=row_num, column=header_map['Søgeord']).value
                match_type = worksheet.cell(row=row_num, column=header_map['Match Type']).value
                
                # Skip empty rows
                if not list_name or not keyword:
                    skipped_rows += 1
                    continue
                
                # Clean and validate data
                list_name = str(list_name).strip()
                category = str(category).strip().lower() if category else 'general'
                keyword = str(keyword).strip()
                match_type = str(match_type).strip().lower() if match_type else 'broad'
                
                # Validate match type
                valid_match_types = ['broad', 'phrase', 'exact']
                if match_type not in valid_match_types:
                    match_type = 'broad'
                
                # Validate category
                valid_categories = [
                    'general', 'job', 'diy', 'competitor', 'location', 
                    'service_specific', 'quality', 'other'
                ]
                if category not in valid_categories:
                    category = 'general'
                
                # Add to lists_data
                if list_name not in lists_data:
                    lists_data[list_name]['category'] = category
                
                lists_data[list_name]['keywords'].append({
                    'text': keyword,
                    'match_type': match_type
                })
                
                processed_rows += 1
                
            except Exception as e:
                errors.append(f'Række {row_num}: {str(e)}')
                skipped_rows += 1
                continue
        
        # Create lists and keywords in database
        created_lists = 0
        created_keywords = 0
        
        for list_name, list_data in lists_data.items():
            try:
                # Get or create the list
                keyword_list, created = NegativeKeywordList.objects.get_or_create(
                    name=list_name,
                    created_by=user,
                    defaults={
                        'category': list_data['category'],
                        'description': f'Importeret fra Excel - {len(list_data["keywords"])} søgeord',
                        'is_active': True,
                        'auto_apply_to_industries': []
                    }
                )
                
                if created:
                    created_lists += 1
                
                # Add keywords to the list
                for keyword_data in list_data['keywords']:
                    # Check if keyword already exists
                    if not NegativeKeyword.objects.filter(
                        keyword_list=keyword_list,
                        keyword_text__iexact=keyword_data['text'],
                        match_type=keyword_data['match_type']
                    ).exists():
                        NegativeKeyword.objects.create(
                            keyword_list=keyword_list,
                            keyword_text=keyword_data['text'],
                            match_type=keyword_data['match_type']
                        )
                        created_keywords += 1
                
            except Exception as e:
                errors.append(f'Liste "{list_name}": {str(e)}')
                continue
        
        # Return results
        return {
            'success': True,
            'summary': {
                'created_lists': created_lists,
                'created_keywords': created_keywords,
                'processed_rows': processed_rows,
                'skipped_rows': skipped_rows,
                'errors': errors[:10]  # Limit to first 10 errors
            }
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Fejl ved læsning af Excel fil: {str(e)}'
        }


# =================================================================
# INDUSTRY MANAGER VIEWS
# =================================================================

def industry_manager(request):
    """Industry Manager - administrer brancher, services og keywords"""
    
    # Get all industries with their related services and keywords
    from .models import IndustryService, ServiceKeyword, IndustryHeadline, NegativeKeywordList
    
    industries = Industry.objects.all().prefetch_related(
        'industry_services__service_keywords',
        'industry_services__service_negative_keyword_lists',
        'industry_headlines'
    ).order_by('name')
    
    # No filtering needed since we use hard delete for services
    
    # Get all active negative keyword lists for service forms
    negative_keyword_lists = NegativeKeywordList.objects.filter(is_active=True).order_by('name')
    
    # Calculate statistics for each industry
    for industry in industries:
        industry.services_total = industry.industry_services.count()
    
    context = {
        'industries': industries,
        'negative_keyword_lists': negative_keyword_lists,
        'total_industries': industries.count(),
        'total_services': IndustryService.objects.count(),
        'total_keywords': ServiceKeyword.objects.count(),
    }
    
    return render(request, 'campaigns/industry_manager.html', context)


def campaign_builder_wizard(request):
    """Campaign Builder Wizard - intelligent kampagne opsætning med multi-step guide"""

    # Hent data for wizard steps
    from .models import BudgetStrategy, AdTemplate
    from usps.models import USPTemplate, USPMainCategory
    from .models import NegativeKeywordList, GeographicRegion, IndustryService, ServiceKeyword

    # Step 1: Industries and Services
    industries = Industry.objects.filter(is_active=True).prefetch_related(
        'industry_services__service_keywords'
    ).order_by('name')

    # Step 2: USPs and Negative Keywords
    usp_categories = USPMainCategory.objects.filter(is_active=True).prefetch_related(
        'usptemplate_set'
    ).order_by('name')

    negative_keyword_lists = NegativeKeywordList.objects.filter(is_active=True).order_by('name')

    # Step 3: Budget Strategies and Ad Templates
    budget_strategies = BudgetStrategy.objects.filter(is_active=True).order_by('-is_default', 'name')
    ad_templates = AdTemplate.objects.filter(is_active=True).order_by('-is_default', 'name')

    # Step 4: Geographic Regions (med prefetched cities)
    geographic_regions = GeographicRegion.objects.filter(is_active=True).prefetch_related('cities').order_by('name')

    # Client selection for Campaign Builder
    clients = Client.objects.all().order_by('name')

    context = {
        'industries': industries,
        'usp_categories': usp_categories,
        'negative_keyword_lists': negative_keyword_lists,
        'budget_strategies': budget_strategies,
        'ad_templates': ad_templates,
        'geographic_regions': geographic_regions,
        'google_maps_api_key': 'AIzaSyBDH6MTS0Hq0ISb0bNQjEAC14321pzM0jw',
        'clients': clients,
    }

    return render(request, 'campaigns/campaign_builder_wizard.html', context)


# =================================================================
# INDUSTRY MANAGER AJAX VIEWS
# =================================================================

@csrf_exempt
def get_industry_services_ajax(request, industry_id):
    """Get services for a specific industry for Campaign Builder"""
    if request.method == 'GET':
        try:
            from .models import IndustryService
            
            industry = Industry.objects.get(id=industry_id)
            services = industry.industry_services.all().order_by('name')
            
            services_data = []
            for service in services:
                keyword_count = service.service_keywords.count()
                services_data.append({
                    'id': service.id,
                    'name': service.name,
                    'description': service.description or '',
                    'keyword_count': keyword_count,
                })
            
            return JsonResponse({
                'success': True,
                'services': services_data,
                'industry_name': industry.name,
                'requires_authorization': industry.requires_authorization
            })
        
        except Industry.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Branche ikke fundet'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@csrf_exempt
def get_service_keywords_ajax(request, service_id):
    """Get keywords for a specific service"""
    if request.method == 'GET':
        try:
            from .models import IndustryService, ServiceKeyword

            service = IndustryService.objects.get(id=service_id)
            keywords = service.service_keywords.all().order_by('-is_primary', 'keyword_text')

            keywords_data = []
            for keyword in keywords:
                keywords_data.append({
                    'id': keyword.id,
                    'keyword_text': keyword.keyword_text,
                    'match_type': keyword.match_type,
                    'match_type_display': keyword.get_match_type_display(),
                    'is_primary': keyword.is_primary,
                    'notes': keyword.notes,
                })

            return JsonResponse({
                'success': True,
                'service': {
                    'id': service.id,
                    'name': service.name,
                    'description': service.description,
                    'color': service.color,
                },
                'keywords': keywords_data,
                'total_keywords': len(keywords_data)
            })

        except IndustryService.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Service ikke fundet'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def get_negative_keyword_lists_ajax(request):
    """Get all negative keyword lists for Campaign Builder"""
    if request.method == 'GET':
        try:
            lists = NegativeKeywordList.objects.all().order_by('name')

            lists_data = []
            for nk_list in lists:
                # Hent tilknyttede service IDs via ManyToMany relation (used_by_services)
                connected_service_ids = list(
                    nk_list.used_by_services.values_list('id', flat=True)
                )

                lists_data.append({
                    'id': nk_list.id,
                    'name': nk_list.name,
                    'description': nk_list.description or '',
                    'keyword_count': nk_list.negative_keywords.count(),
                    'connected_service_ids': connected_service_ids,  # Nye felt
                    'is_city_list': nk_list.name == 'Ekskluderede Byer',  # Marker by-liste
                })

            return JsonResponse({
                'success': True,
                'lists': lists_data
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def add_service_keyword_ajax(request):
    """Add a new keyword to a service"""
    if request.method == 'POST':
        try:
            from .models import IndustryService, ServiceKeyword
            
            service_id = request.POST.get('service_id')
            keyword_text = request.POST.get('keyword_text', '').strip()
            match_type = request.POST.get('match_type', 'phrase')
            is_primary = request.POST.get('is_primary') == 'true'
            notes = request.POST.get('notes', '').strip()
            
            if not service_id or not keyword_text:
                return JsonResponse({'success': False, 'error': 'Service og keyword text er påkrævet'})
            
            service = IndustryService.objects.get(id=service_id)
            
            # Check for duplicates
            if ServiceKeyword.objects.filter(
                service=service,
                keyword_text__iexact=keyword_text,
                match_type=match_type
            ).exists():
                return JsonResponse({'success': False, 'error': f'Keyword "{keyword_text}" med {match_type} match type eksisterer allerede'})
            
            # Create keyword
            keyword = ServiceKeyword.objects.create(
                service=service,
                keyword_text=keyword_text,
                match_type=match_type,
                is_primary=is_primary,
                notes=notes
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Keyword "{keyword_text}" blev tilføjet!',
                'keyword': {
                    'id': keyword.id,
                    'keyword_text': keyword.keyword_text,
                    'match_type': keyword.match_type,
                    'match_type_display': keyword.get_match_type_display(),
                    'is_primary': keyword.is_primary,
                    'notes': keyword.notes,
                }
            })
            
        except IndustryService.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Service ikke fundet'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def update_service_keyword_ajax(request, keyword_id):
    """Update a service keyword"""
    if request.method == 'POST':
        try:
            from .models import ServiceKeyword
            
            keyword = ServiceKeyword.objects.get(id=keyword_id)
            
            keyword_text = request.POST.get('keyword_text', '').strip()
            match_type = request.POST.get('match_type')
            is_primary = request.POST.get('is_primary') == 'true'
            notes = request.POST.get('notes', '').strip()
            
            if not keyword_text:
                return JsonResponse({'success': False, 'error': 'Keyword text er påkrævet'})
            
            # Check for duplicates (excluding current keyword)
            if ServiceKeyword.objects.filter(
                service=keyword.service,
                keyword_text__iexact=keyword_text,
                match_type=match_type
            ).exclude(id=keyword_id).exists():
                return JsonResponse({'success': False, 'error': f'Keyword "{keyword_text}" med {match_type} match type eksisterer allerede'})
            
            # Update keyword
            keyword.keyword_text = keyword_text
            keyword.match_type = match_type
            keyword.is_primary = is_primary
            keyword.notes = notes
            keyword.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Keyword "{keyword_text}" blev opdateret!',
                'keyword': {
                    'id': keyword.id,
                    'keyword_text': keyword.keyword_text,
                    'match_type': keyword.match_type,
                    'match_type_display': keyword.get_match_type_display(),
                    'is_primary': keyword.is_primary,
                    'notes': keyword.notes,
                }
            })
            
        except ServiceKeyword.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Keyword ikke fundet'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def delete_service_keyword_ajax(request, keyword_id):
    """Delete a service keyword"""
    if request.method == 'POST':
        try:
            from .models import ServiceKeyword
            
            keyword = ServiceKeyword.objects.get(id=keyword_id)
            keyword_text = keyword.keyword_text
            
            keyword.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Keyword "{keyword_text}" blev slettet!'
            })
            
        except ServiceKeyword.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Keyword ikke fundet'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def create_industry_ajax(request):
    """Create a new industry"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            synonyms_data = request.POST.get('synonyms', '')
            icon = request.POST.get('icon', '🏢').strip()
            color = request.POST.get('color', '#3B82F6').strip()
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Navn er påkrævet'})
            
            # Parse synonyms (either JSON array or comma separated string)
            synonyms = []
            if synonyms_data:
                try:
                    # Try to parse as JSON first (from new slide panel)
                    synonyms = json.loads(synonyms_data)
                    if not isinstance(synonyms, list):
                        synonyms = []
                except (json.JSONDecodeError, TypeError):
                    # Fall back to comma-separated string parsing (legacy)
                    synonyms = [s.strip() for s in synonyms_data.split(',') if s.strip()]
            
            # Check if industry already exists
            if Industry.objects.filter(name__iexact=name).exists():
                return JsonResponse({'success': False, 'error': f'Branche "{name}" eksisterer allerede'})
            
            # Parse requires_authorization
            requires_authorization = request.POST.get('requires_authorization') == 'true'

            # Create industry
            industry = Industry.objects.create(
                name=name,
                description=description,
                synonyms=synonyms,
                icon=icon,
                color=color,
                is_active=True,
                requires_authorization=requires_authorization
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Branche "{name}" blev oprettet!',
                'industry_id': industry.id
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


# Edit industry and service functions
@csrf_exempt
def edit_industry_ajax(request, industry_id):
    """Edit industry - GET returns current data, POST updates"""
    try:
        industry = Industry.objects.get(id=industry_id)
        
        if request.method == 'GET':
            # Return current industry data for editing
            return JsonResponse({
                'success': True,
                'industry': {
                    'id': industry.id,
                    'name': industry.name,
                    'description': industry.description or '',
                    'synonyms': ', '.join(industry.synonyms) if industry.synonyms else '',
                    'icon': industry.icon or '🏢',
                    'color': industry.color or '#3B82F6',
                    'is_active': industry.is_active,
                    'requires_authorization': industry.requires_authorization
                }
            })
        
        elif request.method == 'POST':
            # Update industry
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            synonyms_data = request.POST.get('synonyms', '')
            icon = request.POST.get('icon', '🏢').strip()
            color = request.POST.get('color', '#3B82F6').strip()
            # Preserve existing is_active if not explicitly sent (slide panel compatibility)
            is_active_field = request.POST.get('is_active')
            is_active = industry.is_active if is_active_field is None else is_active_field == 'true'
            # Preserve existing requires_authorization if not explicitly sent
            requires_auth_field = request.POST.get('requires_authorization')
            requires_authorization = industry.requires_authorization if requires_auth_field is None else requires_auth_field == 'true'

            if not name:
                return JsonResponse({'success': False, 'error': 'Navn er påkrævet'})
            
            # Parse synonyms (either JSON array or comma separated string)
            synonyms = []
            if synonyms_data:
                try:
                    # Try to parse as JSON first (from new slide panel)
                    synonyms = json.loads(synonyms_data)
                    if not isinstance(synonyms, list):
                        synonyms = []
                except (json.JSONDecodeError, TypeError):
                    # Fall back to comma-separated string parsing (legacy)
                    synonyms = [s.strip() for s in synonyms_data.split(',') if s.strip()]
            
            # Check if another industry has this name (exclude current)
            existing = Industry.objects.filter(name__iexact=name).exclude(id=industry_id)
            if existing.exists():
                return JsonResponse({'success': False, 'error': f'Branche "{name}" eksisterer allerede'})
            
            # Update industry
            industry.name = name
            industry.description = description
            industry.synonyms = synonyms
            industry.icon = icon
            industry.color = color
            industry.is_active = is_active
            industry.requires_authorization = requires_authorization
            industry.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Branche "{name}" blev opdateret!'
            })
    
    except Industry.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Branche ikke fundet'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def delete_industry_ajax(request, industry_id):
    """Delete an industry"""
    if request.method == 'DELETE':
        try:
            industry = Industry.objects.get(id=industry_id)
            industry_name = industry.name
            
            # Check if industry has associated services or keywords
            services_count = industry.industry_services.count()
            keywords_count = sum(service.service_keywords.count() for service in industry.industry_services.all())
            
            # Delete the industry (this will cascade delete related services and keywords)
            industry.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Branche "{industry_name}" blev slettet!',
                'deleted_services': services_count,
                'deleted_keywords': keywords_count
            })
            
        except Industry.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Branche ikke fundet'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Fejl ved sletning: {str(e)}'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt 
def edit_service_ajax(request, service_id):
    """Edit service - GET returns current data, POST updates"""
    try:
        service = IndustryService.objects.get(id=service_id)
        
        if request.method == 'GET':
            # Return current service data for editing
            return JsonResponse({
                'success': True,
                'service': {
                    'id': service.id,
                    'name': service.name,
                    'description': service.description or '',
                    'color': service.color or '#8B5CF6',
                    'service_type': service.service_type or 'service',
                    'is_active': service.is_active,
                    'industry_id': service.industry.id,
                    'industry_name': service.industry.name,
                    'negative_keyword_lists': list(service.service_negative_keyword_lists.values_list('id', flat=True))
                }
            })
        
        elif request.method == 'POST':
            # Update service
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            color = request.POST.get('color', '#8B5CF6').strip()
            service_type = request.POST.get('service_type', 'service').strip()
            is_active = request.POST.get('is_active') == 'true'
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Navn er påkrævet'})
            
            # Check if another service in same industry has this name (exclude current)
            existing = IndustryService.objects.filter(
                industry=service.industry, 
                name__iexact=name
            ).exclude(id=service_id)
            
            if existing.exists():
                return JsonResponse({'success': False, 'error': f'Service "{name}" eksisterer allerede i denne branche'})
            
            # Update service
            service.name = name
            service.description = description
            service.color = color
            service.service_type = service_type
            service.is_active = is_active
            service.save()
            
            # Handle negative keyword lists updates
            negative_lists = request.POST.getlist('negative_keyword_lists')
            from .models import NegativeKeywordList
            
            # Clear existing connections
            service.service_negative_keyword_lists.clear()
            
            # Add new connections
            if negative_lists:
                for list_id in negative_lists:
                    try:
                        negative_list = NegativeKeywordList.objects.get(id=list_id, is_active=True)
                        service.service_negative_keyword_lists.add(negative_list)
                    except NegativeKeywordList.DoesNotExist:
                        continue  # Skip invalid list IDs
            
            return JsonResponse({
                'success': True,
                'message': f'Service "{name}" blev opdateret!'
            })
    
    except IndustryService.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Service ikke fundet'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})




@csrf_exempt
def create_service_ajax(request):
    """AJAX endpoint to create new service"""
    if request.method == 'POST':
        try:
            industry_id = request.POST.get('industry_id')
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            color = request.POST.get('color', '#8B5CF6').strip()
            service_type = request.POST.get('service_type', 'service').strip()
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Service navn er påkrævet'})
            
            if not industry_id:
                return JsonResponse({'success': False, 'error': 'Branche ID er påkrævet'})
                
            # Get industry
            industry = get_object_or_404(Industry, id=industry_id)
            
            # Check if service name already exists for this industry
            if IndustryService.objects.filter(
                industry=industry,
                name__iexact=name
            ).exists():
                return JsonResponse({'success': False, 'error': 'En service med dette navn eksisterer allerede for denne branche'})
            
            # Get user for creation - use authenticated user or demo user as fallback  
            if request.user.is_authenticated:
                created_by = request.user
            else:
                # Fallback to demo user for testing
                from django.contrib.auth.models import User
                created_by = User.objects.get(username='demo')
            
            service = IndustryService.objects.create(
                industry=industry,
                name=name,
                description=description,
                color=color,
                service_type=service_type,
                is_active=True,
                created_by=created_by
            )
            
            # Handle negative keyword lists if provided
            negative_lists = request.POST.getlist('negative_keyword_lists')
            if negative_lists:
                from .models import NegativeKeywordList
                for list_id in negative_lists:
                    try:
                        negative_list = NegativeKeywordList.objects.get(id=list_id, is_active=True)
                        service.service_negative_keyword_lists.add(negative_list)
                    except NegativeKeywordList.DoesNotExist:
                        continue  # Skip invalid list IDs
            
            return JsonResponse({
                'success': True,
                'message': f'Service "{name}" blev oprettet!',
                'service': {
                    'id': service.id,
                    'name': service.name,
                    'description': service.description,
                    'color': service.color,
                    'keywords_count': 0
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})




@csrf_exempt
def delete_service_ajax(request, service_id):
    """AJAX endpoint to delete a service"""
    if request.method == 'POST':
        try:
            service = get_object_or_404(IndustryService, id=service_id)
            service_name = service.name
            
            # Hard delete - Django handles CASCADE automatically for:
            # - ServiceKeyword (service keywords)  
            # - ServiceSEOKeyword (SEO keywords)
            # - ServiceNegativeKeywordList (negative keyword connections)
            service.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Service "{service_name}" blev permanent slettet!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def get_service_keywords_ajax(request, service_id):
    """AJAX endpoint to get keywords for a service"""
    if request.method == 'GET':
        try:
            service = get_object_or_404(IndustryService, id=service_id)
            keywords = service.service_keywords.all()
            
            keywords_data = []
            for keyword in keywords:
                keywords_data.append({
                    'id': keyword.id,
                    'keyword_text': keyword.keyword_text,
                    'match_type': keyword.match_type,
                    'match_type_display': keyword.get_match_type_display(),
                    'is_primary': keyword.is_primary,
                    'notes': keyword.notes,
                    'added_at': keyword.added_at.strftime('%d/%m/%Y')
                })
            
            return JsonResponse({
                'success': True,
                'keywords': keywords_data,
                'service_name': service.name
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt  
def add_service_keyword_ajax(request):
    """AJAX endpoint to add keyword to service"""
    if request.method == 'POST':
        try:
            service_id = request.POST.get('service_id')
            keyword_text = request.POST.get('keyword_text', '').strip()
            match_type = request.POST.get('match_type', 'phrase')
            is_primary = request.POST.get('is_primary') == 'true'
            notes = request.POST.get('notes', '').strip()
            
            if not keyword_text:
                return JsonResponse({'success': False, 'error': 'Keyword tekst er påkrævet'})
            
            if not service_id:
                return JsonResponse({'success': False, 'error': 'Service ID er påkrævet'})
                
            # Get service
            service = get_object_or_404(IndustryService, id=service_id)
            
            # Check if keyword with same match type already exists
            if ServiceKeyword.objects.filter(
                service=service,
                keyword_text__iexact=keyword_text,
                match_type=match_type
            ).exists():
                return JsonResponse({
                    'success': False, 
                    'error': f'Keyword "{keyword_text}" med {match_type} match type eksisterer allerede for denne service'
                })
            
            keyword = ServiceKeyword.objects.create(
                service=service,
                keyword_text=keyword_text,
                match_type=match_type,
                is_primary=is_primary,
                notes=notes
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Keyword "{keyword_text}" blev tilføjet!',
                'keyword': {
                    'id': keyword.id,
                    'keyword_text': keyword.keyword_text,
                    'match_type': keyword.match_type,
                    'match_type_display': keyword.get_match_type_display(),
                    'is_primary': keyword.is_primary,
                    'notes': keyword.notes,
                    'added_at': keyword.added_at.strftime('%d/%m/%Y')
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def update_service_keyword_ajax(request, keyword_id):
    """AJAX endpoint to update service keyword"""
    if request.method == 'POST':
        try:
            keyword = get_object_or_404(ServiceKeyword, id=keyword_id)
            
            keyword_text = request.POST.get('keyword_text', '').strip()
            match_type = request.POST.get('match_type', keyword.match_type)
            is_primary = request.POST.get('is_primary') == 'true'
            notes = request.POST.get('notes', '').strip()
            
            if not keyword_text:
                return JsonResponse({'success': False, 'error': 'Keyword tekst er påkrævet'})
            
            # Check if keyword with same match type already exists (excluding current keyword)
            if ServiceKeyword.objects.filter(
                service=keyword.service,
                keyword_text__iexact=keyword_text,
                match_type=match_type
            ).exclude(id=keyword_id).exists():
                return JsonResponse({
                    'success': False, 
                    'error': f'Keyword "{keyword_text}" med {match_type} match type eksisterer allerede for denne service'
                })
            
            keyword.keyword_text = keyword_text
            keyword.match_type = match_type
            keyword.is_primary = is_primary
            keyword.notes = notes
            keyword.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Keyword "{keyword_text}" blev opdateret!',
                'keyword': {
                    'id': keyword.id,
                    'keyword_text': keyword.keyword_text,
                    'match_type': keyword.match_type,
                    'match_type_display': keyword.get_match_type_display(),
                    'is_primary': keyword.is_primary,
                    'notes': keyword.notes
                }
            })
            
        except ServiceKeyword.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Keyword ikke fundet'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def delete_service_keyword_ajax(request, keyword_id):
    """AJAX endpoint to delete service keyword"""
    if request.method == 'POST':
        try:
            keyword = get_object_or_404(ServiceKeyword, id=keyword_id)
            keyword_text = keyword.keyword_text
            
            keyword.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Keyword "{keyword_text}" blev slettet!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})



# =================================================================
# SEO KEYWORDS AJAX VIEWS
# =================================================================

def get_service_seo_keywords_ajax(request, service_id):
    """Get SEO keywords for a specific service"""
    if request.method == "GET":
        try:
            from .models import IndustryService, ServiceSEOKeyword
            
            service = IndustryService.objects.get(id=service_id)
            keywords = service.seo_keywords.all().order_by("-is_primary", "keyword_type", "-search_volume", "keyword_text")
            
            keywords_data = []
            for keyword in keywords:
                keywords_data.append({
                    "id": keyword.id,
                    "keyword_text": keyword.keyword_text,
                    "search_volume": keyword.search_volume,
                    "keyword_type": keyword.keyword_type,
                    "keyword_type_display": keyword.get_keyword_type_display(),
                    "is_primary": keyword.is_primary,
                    "target_url": keyword.target_url,
                    "current_ranking": keyword.current_ranking,
                    "notes": keyword.notes,
                    "added_at": keyword.added_at.strftime("%Y-%m-%d %H:%M"),
                })
            
            return JsonResponse({
                "success": True,
                "service": {
                    "id": service.id,
                    "name": service.name,
                    "description": service.description,
                    "color": service.color,
                },
                "keywords": keywords_data,
                "keywords_count": len(keywords_data),
            })
            
        except IndustryService.DoesNotExist:
            return JsonResponse({"success": False, "error": "Service not found"})
        except Exception as e:
            print(f"Error in get_service_seo_keywords_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


def add_service_seo_keyword_ajax(request, service_id):
    """Add SEO keyword to a service"""
    if request.method == "POST":
        try:
            from .models import IndustryService, ServiceSEOKeyword
            
            service = IndustryService.objects.get(id=service_id)
            
            keyword_text = request.POST.get("keyword_text", "").strip()
            search_volume = request.POST.get("search_volume")
            keyword_type = request.POST.get("keyword_type", "money")
            is_primary = request.POST.get("is_primary") == "true"
            target_url = request.POST.get("target_url", "").strip()
            current_ranking = request.POST.get("current_ranking")
            notes = request.POST.get("notes", "").strip()
            
            if not keyword_text:
                return JsonResponse({"success": False, "error": "Keyword text is required"})
            
            # Check for duplicates
            if ServiceSEOKeyword.objects.filter(service=service, keyword_text__iexact=keyword_text).exists():
                return JsonResponse({"success": False, "error": f"SEO keyword \"{keyword_text}\" already exists for this service"})
            
            # Create keyword
            keyword = ServiceSEOKeyword.objects.create(
                service=service,
                keyword_text=keyword_text,
                search_volume=int(search_volume) if search_volume else None,
                keyword_type=keyword_type,
                is_primary=is_primary,
                target_url=target_url or "",
                current_ranking=int(current_ranking) if current_ranking else None,
                notes=notes
            )
            
            return JsonResponse({
                "success": True,
                "message": f"SEO keyword \"{keyword_text}\" added successfully!",
                "keyword": {
                    "id": keyword.id,
                    "keyword_text": keyword.keyword_text,
                    "search_volume": keyword.search_volume,
                    "keyword_type": keyword.keyword_type,
                    "keyword_type_display": keyword.get_keyword_type_display(),
                    "is_primary": keyword.is_primary,
                }
            })
            
        except IndustryService.DoesNotExist:
            return JsonResponse({"success": False, "error": "Service not found"})
        except ValueError as e:
            return JsonResponse({"success": False, "error": "Invalid number format"})
        except Exception as e:
            print(f"Error in add_service_seo_keyword_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def update_service_seo_keyword_ajax(request, keyword_id):
    """Update SEO keyword for a service"""
    if request.method == "POST":
        try:
            from .models import ServiceSEOKeyword
            
            keyword = ServiceSEOKeyword.objects.get(id=keyword_id)
            
            keyword_text = request.POST.get("keyword_text", "").strip()
            search_volume = request.POST.get("search_volume")
            keyword_type = request.POST.get("keyword_type", "money")
            is_primary = request.POST.get("is_primary") == "true"
            target_url = request.POST.get("target_url", "").strip()
            current_ranking = request.POST.get("current_ranking")
            notes = request.POST.get("notes", "").strip()
            
            if not keyword_text:
                return JsonResponse({"success": False, "error": "Keyword text is required"})
            
            # Check for duplicates (excluding current keyword)
            if ServiceSEOKeyword.objects.filter(
                service=keyword.service, 
                keyword_text__iexact=keyword_text
            ).exclude(id=keyword_id).exists():
                return JsonResponse({"success": False, "error": f"SEO keyword \"{keyword_text}\" already exists for this service"})
            
            # Update keyword
            keyword.keyword_text = keyword_text
            keyword.search_volume = int(search_volume) if search_volume else None
            keyword.keyword_type = keyword_type
            keyword.is_primary = is_primary
            keyword.target_url = target_url or ""
            keyword.current_ranking = int(current_ranking) if current_ranking else None
            keyword.notes = notes
            keyword.save()
            
            return JsonResponse({
                "success": True,
                "message": f"SEO keyword \"{keyword_text}\" updated successfully!",
                "keyword": {
                    "id": keyword.id,
                    "keyword_text": keyword.keyword_text,
                    "search_volume": keyword.search_volume,
                    "keyword_type": keyword.keyword_type,
                    "keyword_type_display": keyword.get_keyword_type_display(),
                    "is_primary": keyword.is_primary,
                }
            })
            
        except ServiceSEOKeyword.DoesNotExist:
            return JsonResponse({"success": False, "error": "SEO keyword not found"})
        except ValueError as e:
            return JsonResponse({"success": False, "error": "Invalid number format"})
        except Exception as e:
            print(f"Error in update_service_seo_keyword_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def delete_service_seo_keyword_ajax(request, keyword_id):
    """Delete SEO keyword from a service"""
    if request.method == "POST":
        try:
            from .models import ServiceSEOKeyword

            keyword = ServiceSEOKeyword.objects.get(id=keyword_id)
            keyword_text = keyword.keyword_text

            keyword.delete()

            return JsonResponse({
                "success": True,
                "message": f"SEO keyword \"{keyword_text}\" deleted successfully!"
            })

        except ServiceSEOKeyword.DoesNotExist:
            return JsonResponse({"success": False, "error": "SEO keyword not found"})
        except Exception as e:
            print(f"Error in delete_service_seo_keyword_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


# =================================================================
# SERVICE META TAG EXAMPLES (Few-Shot AI Learning)
# =================================================================

def get_service_meta_examples_ajax(request, service_id):
    """Get meta tag examples for a specific service"""
    if request.method == "GET":
        try:
            from .models import IndustryService, ServiceMetaExample

            service = IndustryService.objects.get(id=service_id)
            examples = service.meta_examples.all()

            examples_data = []
            for ex in examples:
                examples_data.append({
                    "id": ex.id,
                    "meta_title": ex.meta_title,
                    "meta_description": ex.meta_description,
                    "order": ex.order,
                    "created_at": ex.created_at.strftime("%Y-%m-%d %H:%M"),
                })

            return JsonResponse({
                "success": True,
                "service": {
                    "id": service.id,
                    "name": service.name,
                },
                "examples": examples_data,
                "examples_count": len(examples_data),
            })

        except IndustryService.DoesNotExist:
            return JsonResponse({"success": False, "error": "Service not found"})
        except Exception as e:
            print(f"Error in get_service_meta_examples_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


def add_service_meta_example_ajax(request, service_id):
    """Add meta tag example to a service (no limit)"""
    if request.method == "POST":
        try:
            from .models import IndustryService, ServiceMetaExample
            import json

            service = IndustryService.objects.get(id=service_id)

            # Support both form data and JSON
            if request.content_type == 'application/json':
                data = json.loads(request.body)
                meta_title = data.get("meta_title", "").strip()
                meta_description = data.get("meta_description", "").strip()
            else:
                meta_title = request.POST.get("meta_title", "").strip()
                meta_description = request.POST.get("meta_description", "").strip()

            if not meta_title or not meta_description:
                return JsonResponse({"success": False, "error": "Both meta title and description are required"})

            # Get next order number
            from django.db.models import Max
            max_order = service.meta_examples.aggregate(Max('order'))['order__max'] or 0

            # Create example
            example = ServiceMetaExample.objects.create(
                service=service,
                meta_title=meta_title,
                meta_description=meta_description,
                order=max_order + 1
            )

            return JsonResponse({
                "success": True,
                "example": {
                    "id": example.id,
                    "meta_title": example.meta_title,
                    "meta_description": example.meta_description,
                    "order": example.order,
                },
                "message": "Meta tag example added successfully!"
            })

        except IndustryService.DoesNotExist:
            return JsonResponse({"success": False, "error": "Service not found"})
        except Exception as e:
            print(f"Error in add_service_meta_example_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


def update_service_meta_example_ajax(request, example_id):
    """Update existing meta tag example"""
    if request.method == "POST":
        try:
            from .models import ServiceMetaExample
            import json

            example = ServiceMetaExample.objects.get(id=example_id)

            # Support both form data and JSON
            if request.content_type == 'application/json':
                data = json.loads(request.body)
                meta_title = data.get("meta_title", "").strip()
                meta_description = data.get("meta_description", "").strip()
            else:
                meta_title = request.POST.get("meta_title", "").strip()
                meta_description = request.POST.get("meta_description", "").strip()

            if meta_title:
                example.meta_title = meta_title
            if meta_description:
                example.meta_description = meta_description

            example.save()

            return JsonResponse({
                "success": True,
                "example": {
                    "id": example.id,
                    "meta_title": example.meta_title,
                    "meta_description": example.meta_description,
                },
                "message": "Meta tag example updated successfully!"
            })

        except ServiceMetaExample.DoesNotExist:
            return JsonResponse({"success": False, "error": "Meta example not found"})
        except Exception as e:
            print(f"Error in update_service_meta_example_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


def delete_service_meta_example_ajax(request, example_id):
    """Delete meta tag example"""
    if request.method == "POST":
        try:
            from .models import ServiceMetaExample

            example = ServiceMetaExample.objects.get(id=example_id)
            example.delete()

            return JsonResponse({
                "success": True,
                "message": "Meta tag example deleted successfully!"
            })

        except ServiceMetaExample.DoesNotExist:
            return JsonResponse({"success": False, "error": "Meta example not found"})
        except Exception as e:
            print(f"Error in delete_service_meta_example_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


# =================================================================
# NEGATIVE KEYWORDS INTEGRATION VIEWS
# =================================================================

@csrf_exempt
def search_negative_keyword_lists_ajax(request):
    """Search available negative keyword lists"""
    if request.method == "GET":
        try:
            from .models import NegativeKeywordList
            
            query = request.GET.get('query', '').strip()
            industry_id = request.GET.get('industry_id')
            
            # Base queryset - only active lists
            lists = NegativeKeywordList.objects.filter(is_active=True)
            
            # Filter by query if provided
            if query:
                from django.db import models as django_models
                lists = lists.filter(
                    django_models.Q(name__icontains=query) |
                    django_models.Q(description__icontains=query) |
                    django_models.Q(category__icontains=query)
                )
            
            # Filter by industry if provided
            if industry_id:
                from .models import Industry
                try:
                    industry = Industry.objects.get(id=industry_id)
                    lists = lists.filter(
                        django_models.Q(industry=industry) |
                        django_models.Q(auto_apply_to_industries__contains=[industry_id]) |
                        django_models.Q(industry__isnull=True)  # Include general lists
                    )
                except Industry.DoesNotExist:
                    pass
            
            # Prepare response data
            lists_data = []
            for keyword_list in lists.order_by('name'):
                lists_data.append({
                    'id': keyword_list.id,
                    'name': keyword_list.name,
                    'category': keyword_list.category,
                    'category_display': keyword_list.get_category_display(),
                    'description': keyword_list.description or '',
                    'keywords_count': keyword_list.keywords_count,
                    'icon': keyword_list.icon or '📋',
                    'color': keyword_list.color or '#8B5CF6',
                    'industry_name': keyword_list.industry.name if keyword_list.industry else 'Alle brancher',
                    'created_at': keyword_list.created_at.strftime('%Y-%m-%d')
                })
            
            return JsonResponse({
                'success': True,
                'lists': lists_data,
                'total_count': len(lists_data),
                'query': query
            })
            
        except Exception as e:
            print(f"Error in search_negative_keyword_lists_ajax: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def get_service_negative_lists_ajax(request, service_id):
    """Get negative keyword lists connected to a service"""
    if request.method == "GET":
        try:
            from .models import IndustryService, ServiceNegativeKeywordList
            
            service = IndustryService.objects.get(id=service_id)
            
            # Get connected lists through the junction table
            connections = ServiceNegativeKeywordList.objects.filter(
                service=service, 
                is_active=True
            ).select_related('negative_list')
            
            lists_data = []
            for connection in connections:
                keyword_list = connection.negative_list
                lists_data.append({
                    'connection_id': connection.id,
                    'list_id': keyword_list.id,
                    'name': keyword_list.name,
                    'category': keyword_list.category,
                    'category_display': keyword_list.get_category_display(),
                    'description': keyword_list.description or '',
                    'keywords_count': keyword_list.keywords_count,
                    'icon': keyword_list.icon or '📋',
                    'color': keyword_list.color or '#8B5CF6',
                    'connected_at': connection.connected_at.strftime('%Y-%m-%d'),
                })
            
            return JsonResponse({
                'success': True,
                'service': {
                    'id': service.id,
                    'name': service.name,
                    'industry_name': service.industry.name
                },
                'lists': lists_data,
                'lists_count': len(lists_data)
            })
            
        except IndustryService.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Service not found'})
        except Exception as e:
            print(f"Error in get_service_negative_lists_ajax: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def connect_negative_list_to_service_ajax(request, service_id):
    """Connect a negative keyword list to a service"""
    if request.method == "POST":
        try:
            from .models import IndustryService, NegativeKeywordList, ServiceNegativeKeywordList
            
            service = IndustryService.objects.get(id=service_id)
            list_id = request.POST.get('list_id')
            
            if not list_id:
                return JsonResponse({'success': False, 'error': 'List ID is required'})
            
            negative_list = NegativeKeywordList.objects.get(id=list_id)
            
            # Check if already connected
            connection, created = ServiceNegativeKeywordList.objects.get_or_create(
                service=service,
                negative_list=negative_list,
                defaults={'is_active': True}
            )
            
            if created:
                message = f'Negative keyword liste "{negative_list.name}" blev tilkoblet til "{service.name}"'
            else:
                # Reactivate if was disabled
                connection.is_active = True
                connection.save()
                message = f'Negative keyword liste "{negative_list.name}" blev genaktiveret for "{service.name}"'
            
            return JsonResponse({
                'success': True,
                'message': message,
                'connection': {
                    'connection_id': connection.id,
                    'list_id': negative_list.id,
                    'name': negative_list.name,
                    'category': negative_list.category,
                    'keywords_count': negative_list.keywords_count,
                    'icon': negative_list.icon,
                    'color': negative_list.color
                }
            })
            
        except IndustryService.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Service not found'})
        except NegativeKeywordList.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Negative keyword list not found'})
        except Exception as e:
            print(f"Error in connect_negative_list_to_service_ajax: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def disconnect_negative_list_from_service_ajax(request, connection_id):
    """Disconnect a negative keyword list from a service"""
    if request.method == "POST":
        try:
            from .models import ServiceNegativeKeywordList
            
            connection = ServiceNegativeKeywordList.objects.get(id=connection_id)
            list_name = connection.negative_list.name
            service_name = connection.service.name
            
            # Soft delete by setting is_active to False
            connection.is_active = False
            connection.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Negative keyword liste "{list_name}" blev frakoblet fra "{service_name}"'
            })
            
        except ServiceNegativeKeywordList.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Connection not found'})
        except Exception as e:
            print(f"Error in disconnect_negative_list_from_service_ajax: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})



# =================================================================
# INDUSTRY KEYWORDS MANAGEMENT VIEWS  
# =================================================================

@csrf_exempt
def get_industry_keywords_ajax(request, industry_id):
    """Get Google Ads keywords for an industry"""
    if request.method == "GET":
        try:
            from .models import Industry, IndustryKeyword
            
            industry = Industry.objects.get(id=industry_id)
            keywords = IndustryKeyword.objects.filter(industry=industry).order_by('-is_primary', 'keyword_text')
            
            keywords_data = []
            for keyword in keywords:
                keywords_data.append({
                    "id": keyword.id,
                    "keyword_text": keyword.keyword_text,
                    "match_type": keyword.match_type,
                    "match_type_display": keyword.get_match_type_display(),
                    "is_primary": keyword.is_primary,
                    "inherited_by_services": keyword.inherited_by_services,
                    "notes": keyword.notes,
                    "added_at": keyword.added_at.strftime('%Y-%m-%d')
                })
            
            return JsonResponse({
                "success": True,
                "industry": {
                    "id": industry.id,
                    "name": industry.name
                },
                "keywords": keywords_data,
                "keywords_count": len(keywords_data)
            })
            
        except Industry.DoesNotExist:
            return JsonResponse({"success": False, "error": "Industry not found"})
        except Exception as e:
            print(f"Error in get_industry_keywords_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def add_industry_keyword_ajax(request, industry_id):
    """Add Google Ads keyword to an industry"""
    if request.method == "POST":
        try:
            from .models import Industry, IndustryKeyword
            
            industry = Industry.objects.get(id=industry_id)
            
            keyword_text = request.POST.get("keyword_text", "").strip()
            match_type = request.POST.get("match_type", "phrase")
            is_primary = request.POST.get("is_primary") == "true"
            inherited_by_services = request.POST.get("inherited_by_services", "true") == "true"
            notes = request.POST.get("notes", "").strip()
            
            if not keyword_text:
                return JsonResponse({"success": False, "error": "Keyword text is required"})
            
            # Check for duplicates
            if IndustryKeyword.objects.filter(industry=industry, keyword_text__iexact=keyword_text, match_type=match_type).exists():
                return JsonResponse({"success": False, "error": f"Google Ads keyword \"{keyword_text}\" with {match_type} match type already exists for this industry"})
            
            # Create keyword
            keyword = IndustryKeyword.objects.create(
                industry=industry,
                keyword_text=keyword_text,
                match_type=match_type,
                is_primary=is_primary,
                inherited_by_services=inherited_by_services,
                notes=notes
            )
            
            return JsonResponse({
                "success": True,
                "message": f"Google Ads keyword \"{keyword_text}\" added successfully to industry!",
                "keyword": {
                    "id": keyword.id,
                    "keyword_text": keyword.keyword_text,
                    "match_type": keyword.match_type,
                    "match_type_display": keyword.get_match_type_display(),
                    "is_primary": keyword.is_primary,
                    "inherited_by_services": keyword.inherited_by_services
                }
            })
            
        except Industry.DoesNotExist:
            return JsonResponse({"success": False, "error": "Industry not found"})
        except Exception as e:
            print(f"Error in add_industry_keyword_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def update_industry_keyword_ajax(request, keyword_id):
    """Update an industry Google Ads keyword"""
    if request.method == "POST":
        try:
            from .models import IndustryKeyword
            
            keyword = IndustryKeyword.objects.get(id=keyword_id)
            
            keyword_text = request.POST.get("keyword_text", "").strip()
            match_type = request.POST.get("match_type", keyword.match_type)
            is_primary = request.POST.get("is_primary") == "true"
            inherited_by_services = request.POST.get("inherited_by_services", str(keyword.inherited_by_services).lower()) == "true"
            
            if not keyword_text:
                return JsonResponse({"success": False, "error": "Keyword text is required"})
            
            # Check for duplicates (excluding current keyword)
            if IndustryKeyword.objects.filter(
                industry=keyword.industry, 
                keyword_text__iexact=keyword_text, 
                match_type=match_type
            ).exclude(id=keyword_id).exists():
                return JsonResponse({"success": False, "error": f"Google Ads keyword \"{keyword_text}\" with {match_type} match type already exists for this industry"})
            
            # Update keyword
            keyword.keyword_text = keyword_text
            keyword.match_type = match_type
            keyword.is_primary = is_primary
            keyword.inherited_by_services = inherited_by_services
            keyword.save()
            
            return JsonResponse({
                "success": True,
                "message": f"Google Ads keyword \"{keyword_text}\" updated successfully!",
                "keyword": {
                    "id": keyword.id,
                    "keyword_text": keyword.keyword_text,
                    "match_type": keyword.match_type,
                    "match_type_display": keyword.get_match_type_display(),
                    "is_primary": keyword.is_primary,
                    "inherited_by_services": keyword.inherited_by_services
                }
            })
            
        except IndustryKeyword.DoesNotExist:
            return JsonResponse({"success": False, "error": "Keyword not found"})
        except Exception as e:
            print(f"Error in update_industry_keyword_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def delete_industry_keyword_ajax(request, keyword_id):
    """Delete an industry Google Ads keyword"""
    if request.method == "POST":
        try:
            from .models import IndustryKeyword
            
            keyword = IndustryKeyword.objects.get(id=keyword_id)
            keyword_text = keyword.keyword_text
            industry_name = keyword.industry.name
            
            keyword.delete()
            
            return JsonResponse({
                "success": True,
                "message": f"Google Ads keyword \"{keyword_text}\" deleted successfully from industry \"{industry_name}\"!"
            })
            
        except IndustryKeyword.DoesNotExist:
            return JsonResponse({"success": False, "error": "Keyword not found"})
        except Exception as e:
            print(f"Error in delete_industry_keyword_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def get_industry_seo_keywords_ajax(request, industry_id):
    """Get SEO keywords for an industry"""
    if request.method == "GET":
        try:
            from .models import Industry, IndustrySEOKeyword
            
            industry = Industry.objects.get(id=industry_id)
            keywords = IndustrySEOKeyword.objects.filter(industry=industry).order_by('-is_primary', 'search_volume', 'keyword_text')
            
            keywords_data = []
            for keyword in keywords:
                keywords_data.append({
                    "id": keyword.id,
                    "keyword_text": keyword.keyword_text,
                    "search_volume": keyword.search_volume,
                    "keyword_type": keyword.keyword_type,
                    "keyword_type_display": keyword.get_keyword_type_display(),
                    "is_primary": keyword.is_primary,
                    "inherited_by_services": keyword.inherited_by_services,
                    "notes": keyword.notes,
                    "added_at": keyword.added_at.strftime('%Y-%m-%d')
                })
            
            return JsonResponse({
                "success": True,
                "industry": {
                    "id": industry.id,
                    "name": industry.name
                },
                "keywords": keywords_data,
                "keywords_count": len(keywords_data)
            })
            
        except Industry.DoesNotExist:
            return JsonResponse({"success": False, "error": "Industry not found"})
        except Exception as e:
            print(f"Error in get_industry_seo_keywords_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def add_industry_seo_keyword_ajax(request, industry_id):
    """Add SEO keyword to an industry"""
    if request.method == "POST":
        try:
            from .models import Industry, IndustrySEOKeyword
            
            industry = Industry.objects.get(id=industry_id)
            
            keyword_text = request.POST.get("keyword_text", "").strip()
            search_volume = request.POST.get("search_volume")
            keyword_type = request.POST.get("keyword_type", "money")
            is_primary = request.POST.get("is_primary") == "true"
            inherited_by_services = request.POST.get("inherited_by_services", "true") == "true"
            notes = request.POST.get("notes", "").strip()
            
            if not keyword_text:
                return JsonResponse({"success": False, "error": "Keyword text is required"})
            
            # Check for duplicates
            if IndustrySEOKeyword.objects.filter(industry=industry, keyword_text__iexact=keyword_text).exists():
                return JsonResponse({"success": False, "error": f"SEO keyword \"{keyword_text}\" already exists for this industry"})
            
            # Create keyword
            keyword = IndustrySEOKeyword.objects.create(
                industry=industry,
                keyword_text=keyword_text,
                search_volume=int(search_volume) if search_volume else None,
                keyword_type=keyword_type,
                is_primary=is_primary,
                inherited_by_services=inherited_by_services,
                notes=notes
            )
            
            return JsonResponse({
                "success": True,
                "message": f"SEO keyword \"{keyword_text}\" added successfully to industry!",
                "keyword": {
                    "id": keyword.id,
                    "keyword_text": keyword.keyword_text,
                    "search_volume": keyword.search_volume,
                    "keyword_type": keyword.keyword_type,
                    "keyword_type_display": keyword.get_keyword_type_display(),
                    "is_primary": keyword.is_primary,
                    "inherited_by_services": keyword.inherited_by_services
                }
            })
            
        except Industry.DoesNotExist:
            return JsonResponse({"success": False, "error": "Industry not found"})
        except Exception as e:
            print(f"Error in add_industry_seo_keyword_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def update_industry_seo_keyword_ajax(request, keyword_id):
    """Update an industry SEO keyword"""
    if request.method == "POST":
        try:
            from .models import IndustrySEOKeyword
            
            keyword = IndustrySEOKeyword.objects.get(id=keyword_id)
            
            keyword_text = request.POST.get("keyword_text", "").strip()
            search_volume = request.POST.get("search_volume")
            keyword_type = request.POST.get("keyword_type", keyword.keyword_type)
            is_primary = request.POST.get("is_primary") == "true"
            inherited_by_services = request.POST.get("inherited_by_services", str(keyword.inherited_by_services).lower()) == "true"
            
            if not keyword_text:
                return JsonResponse({"success": False, "error": "Keyword text is required"})
            
            # Check for duplicates (excluding current keyword)
            if IndustrySEOKeyword.objects.filter(
                industry=keyword.industry, 
                keyword_text__iexact=keyword_text
            ).exclude(id=keyword_id).exists():
                return JsonResponse({"success": False, "error": f"SEO keyword \"{keyword_text}\" already exists for this industry"})
            
            # Update keyword
            keyword.keyword_text = keyword_text
            keyword.search_volume = int(search_volume) if search_volume else None
            keyword.keyword_type = keyword_type
            keyword.is_primary = is_primary
            keyword.inherited_by_services = inherited_by_services
            keyword.save()
            
            return JsonResponse({
                "success": True,
                "message": f"SEO keyword \"{keyword_text}\" updated successfully!",
                "keyword": {
                    "id": keyword.id,
                    "keyword_text": keyword.keyword_text,
                    "search_volume": keyword.search_volume,
                    "keyword_type": keyword.keyword_type,
                    "keyword_type_display": keyword.get_keyword_type_display(),
                    "is_primary": keyword.is_primary,
                    "inherited_by_services": keyword.inherited_by_services
                }
            })
            
        except IndustrySEOKeyword.DoesNotExist:
            return JsonResponse({"success": False, "error": "Keyword not found"})
        except Exception as e:
            print(f"Error in update_industry_seo_keyword_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def delete_industry_seo_keyword_ajax(request, keyword_id):
    """Delete an industry SEO keyword"""
    if request.method == "POST":
        try:
            from .models import IndustrySEOKeyword
            
            keyword = IndustrySEOKeyword.objects.get(id=keyword_id)
            keyword_text = keyword.keyword_text
            industry_name = keyword.industry.name
            
            keyword.delete()
            
            return JsonResponse({
                "success": True,
                "message": f"SEO keyword \"{keyword_text}\" deleted successfully from industry \"{industry_name}\"!"
            })
            
        except IndustrySEOKeyword.DoesNotExist:
            return JsonResponse({"success": False, "error": "Keyword not found"})
        except Exception as e:
            print(f"Error in delete_industry_seo_keyword_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def get_service_negative_lists_ajax(request, service_id):
    """Get negative keyword lists for a specific service"""
    if request.method == "GET":
        try:
            from .models import IndustryService, NegativeKeywordList
            
            service = IndustryService.objects.get(id=service_id)
            
            # Get all available negative keyword lists
            available_lists = NegativeKeywordList.objects.filter(is_active=True).order_by('name')
            
            # Get connected negative keyword lists for this service
            connected_lists = []
            for connection in service.service_negative_keyword_lists.all():
                connected_lists.append({
                    'list_id': connection.id,
                    'connection_id': connection.id,  # For M2M, this would be different
                    'name': connection.name,
                    'description': connection.description,
                    'keywords_count': connection.negative_keywords.count(),
                    'category': connection.category,
                    'category_display': connection.get_category_display(),
                    'connected_at': connection.created_at.strftime('%d/%m/%Y') if hasattr(connection, 'created_at') else 'N/A'
                })
            
            # Prepare available lists data
            available_lists_data = []
            for neg_list in available_lists:
                available_lists_data.append({
                    'id': neg_list.id,
                    'name': neg_list.name,
                    'description': neg_list.description,
                    'keywords_count': neg_list.negative_keywords.count(),
                    'category': neg_list.category,
                    'category_display': neg_list.get_category_display(),
                    'icon': getattr(neg_list, 'icon', '📋'),
                    'color': getattr(neg_list, 'color', '#8B5CF6'),
                    'industry_name': getattr(neg_list.industry, 'name', 'Generel') if hasattr(neg_list, 'industry') and neg_list.industry else 'Generel'
                })
            
            return JsonResponse({
                "success": True,
                "service_name": service.name,
                "connected_lists": connected_lists,
                "available_lists": available_lists_data
            })
            
        except IndustryService.DoesNotExist:
            return JsonResponse({"success": False, "error": "Service not found"})
        except Exception as e:
            print(f"Error in get_service_negative_lists_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def connect_negative_list_to_service_ajax(request, service_id):
    """Connect a negative keyword list to a service"""
    if request.method == "POST":
        try:
            from .models import IndustryService, NegativeKeywordList
            
            service = IndustryService.objects.get(id=service_id)
            list_id = request.POST.get('list_id')
            
            if not list_id:
                return JsonResponse({"success": False, "error": "List ID is required"})
            
            negative_list = NegativeKeywordList.objects.get(id=list_id)
            
            # Check if already connected
            if service.service_negative_keyword_lists.filter(id=list_id).exists():
                return JsonResponse({"success": False, "error": f"Liste '{negative_list.name}' er allerede tilknyttet denne service"})
            
            # Connect the list
            service.service_negative_keyword_lists.add(negative_list)
            
            return JsonResponse({
                "success": True,
                "message": f"Negative keyword liste '{negative_list.name}' tilknyttet succesfuldt til '{service.name}'"
            })
            
        except IndustryService.DoesNotExist:
            return JsonResponse({"success": False, "error": "Service not found"})
        except NegativeKeywordList.DoesNotExist:
            return JsonResponse({"success": False, "error": "Negative keyword list not found"})
        except Exception as e:
            print(f"Error in connect_negative_list_to_service_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def disconnect_negative_list_from_service_ajax(request, connection_id):
    """Disconnect a negative keyword list from a service"""
    if request.method == "POST":
        try:
            from .models import IndustryService, NegativeKeywordList
            
            service_id = request.POST.get('service_id')
            if not service_id:
                return JsonResponse({"success": False, "error": "Service ID is required"})
            
            service = IndustryService.objects.get(id=service_id)
            
            # For M2M relationship, connection_id is actually the list_id
            negative_list = NegativeKeywordList.objects.get(id=connection_id)
            
            # Check if connected
            if not service.service_negative_keyword_lists.filter(id=connection_id).exists():
                return JsonResponse({"success": False, "error": f"Liste '{negative_list.name}' er ikke tilknyttet denne service"})
            
            # Disconnect the list
            service.service_negative_keyword_lists.remove(negative_list)
            
            return JsonResponse({
                "success": True,
                "message": f"Negative keyword liste '{negative_list.name}' fjernet fra '{service.name}'"
            })
            
        except IndustryService.DoesNotExist:
            return JsonResponse({"success": False, "error": "Service not found"})
        except NegativeKeywordList.DoesNotExist:
            return JsonResponse({"success": False, "error": "Negative keyword list not found"})
        except Exception as e:
            print(f"Error in disconnect_negative_list_from_service_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def search_negative_keyword_lists_ajax(request):
    """Search negative keyword lists"""
    if request.method == "GET":
        try:
            from django.db import models
            from .models import NegativeKeywordList
            
            query = request.GET.get('q', '').strip()
            
            if not query or len(query) < 2:
                return JsonResponse({
                    "success": True,
                    "lists": []
                })
            
            # Search in name, description and category
            lists = NegativeKeywordList.objects.filter(
                models.Q(name__icontains=query) |
                models.Q(description__icontains=query) |
                models.Q(category__icontains=query),
                is_active=True
            ).order_by('name')[:20]  # Limit to 20 results
            
            lists_data = []
            for neg_list in lists:
                lists_data.append({
                    'id': neg_list.id,
                    'name': neg_list.name,
                    'description': neg_list.description,
                    'keywords_count': neg_list.negative_keywords.count(),
                    'category': neg_list.category,
                    'category_display': neg_list.get_category_display(),
                    'icon': getattr(neg_list, 'icon', '📋'),
                    'color': getattr(neg_list, 'color', '#8B5CF6'),
                    'industry_name': getattr(neg_list.industry, 'name', 'Generel') if hasattr(neg_list, 'industry') and neg_list.industry else 'Generel'
                })
            
            return JsonResponse({
                "success": True,
                "lists": lists_data
            })
            
        except Exception as e:
            print(f"Error in search_negative_keyword_lists_ajax: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


@csrf_exempt
def generate_descriptions_ajax(request):
    """AJAX endpoint til AI-genererede beskrivelser."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        data = json.loads(request.body)
        service_name = data.get('service_name', '')
        industry_name = data.get('industry_name', '')
        usps = data.get('usps', [])
        keywords = data.get('keywords', [])

        from ai_integration.services import DescriptionGenerator
        generator = DescriptionGenerator()
        descriptions = generator.generate_descriptions(
            service_name, industry_name, usps, keywords
        )

        return JsonResponse({
            'success': True,
            'descriptions': descriptions
        })

    except ValueError as e:
        # API key not configured
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
def generate_company_description_ajax(request):
    """AJAX endpoint til AI-genereret virksomhedsbeskrivelse med Perplexity research."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        data = json.loads(request.body)
        website_url = data.get('website_url', '')
        industries = data.get('industries', [])
        services = data.get('services', [])
        usps = data.get('usps', [])
        geographic_areas = data.get('geographic_areas', [])

        # Step 0: Scrape the website directly (get actual content)
        website_content = ""
        try:
            from ai_integration.services import WebsiteScraper
            scraper = WebsiteScraper()
            website_content = scraper.scrape_website(website_url)
            if website_content:
                print(f"Website scraping success: {len(website_content)} chars")
        except Exception as e:
            # Scraping failed - continue without it
            print(f"Website scraping failed: {e}")

        # Step 1: Research with Perplexity (if API key is configured)
        online_research = ""
        try:
            from ai_integration.services import PerplexityResearcher
            researcher = PerplexityResearcher()
            online_research = researcher.research_company(website_url, industries, services)
        except ValueError:
            # Perplexity not configured - continue without research
            pass
        except Exception as e:
            # Research failed - continue without it
            print(f"Perplexity research step failed: {e}")

        # Step 2: Generate description with GPT-4
        from ai_integration.services import DescriptionGenerator
        generator = DescriptionGenerator()
        result = generator.generate_company_description(
            website_url=website_url,
            industries=industries,
            services=services,
            usps=usps,
            geographic_areas=geographic_areas,
            online_research=online_research,
            website_content=website_content
        )

        # Fix any encoding issues (UTF-8 decoded as Latin-1)
        def fix_encoding(text):
            if not text:
                return text
            try:
                # Check for mojibake pattern (UTF-8 decoded as Latin-1)
                if 'Ã' in text:
                    # Re-encode as Latin-1 and decode as UTF-8
                    return text.encode('latin-1').decode('utf-8')
            except (UnicodeDecodeError, UnicodeEncodeError):
                pass
            return text

        description = fix_encoding(result['description'])
        key_points = [fix_encoding(kp) for kp in result.get('key_points', [])]

        return JsonResponse({
            'success': True,
            'description': description,
            'key_points': key_points,
            'profile': result.get('profile', {}),
            'used_research': bool(online_research),
            'model_used': result.get('model_used', 'unknown')
        }, json_dumps_params={'ensure_ascii': False})

    except ValueError as e:
        # API key not configured
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, json_dumps_params={'ensure_ascii': False})


@csrf_exempt
def analyze_website_for_usps_ajax(request):
    """
    AJAX endpoint til at analysere hjemmeside-indhold og matche mod USP templates.
    Bruges til smart pre-fill af USPs baseret på hjemmesidens indhold.

    Accepts:
        - website_url: URL to analyze
        - industry_ids: List of industry IDs to filter USP templates
        - scrape_mode: Number of pages to scrape (10, 50, 100, or 0 for all)
        - client_id: Optional client ID for permanent storage
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        data = json.loads(request.body)
        website_url = data.get('website_url', '')
        industry_ids = data.get('industry_ids', [])
        scrape_mode = data.get('scrape_mode', 10)  # Default: 10 pages
        client_id = data.get('client_id', None)  # Optional: for permanent storage

        if not website_url:
            return JsonResponse({
                'success': True,
                'analysis': {
                    'matched_usps': [],
                    'custom_usps': [],
                    'extracted_facts': {}
                },
                'scraped_at': None
            })

        # Step 1: Scrape website using ComprehensiveWebsiteScraper (multi-page)
        from ai_integration.services import ComprehensiveWebsiteScraper
        scraper = ComprehensiveWebsiteScraper()

        # Convert scrape_mode to max_pages (0 means None/all)
        max_pages = scrape_mode if scrape_mode > 0 else None

        scrape_result = scraper.scrape_website(
            url=website_url,
            max_pages=max_pages,
            client_id=client_id
        )

        if not scrape_result or not scrape_result.get('combined_content'):
            return JsonResponse({
                'success': True,
                'analysis': {
                    'matched_usps': [],
                    'custom_usps': [],
                    'extracted_facts': {}
                },
                'scraped_at': None,
                'message': 'Could not scrape website'
            })

        # Use combined content from all scraped pages
        website_content = scrape_result['combined_content']

        # Step 2: Get relevant USP templates
        from usps.models import USPTemplate
        from django.db.models import Q

        usp_templates = USPTemplate.objects.filter(is_active=True)
        if industry_ids:
            usp_templates = usp_templates.filter(
                Q(ideal_for_industries__id__in=industry_ids) |
                Q(ideal_for_industries__isnull=True)
            ).distinct()

        # Step 3: Format templates for AI prompt
        templates_formatted = []
        for template in usp_templates[:60]:  # Limit to prevent token overflow
            templates_formatted.append({
                'id': template.id,
                'text': template.text,
                'keywords': template.keywords if hasattr(template, 'keywords') else '',
                'category': template.main_category.name if template.main_category else 'Andet'
            })

        # Step 4: Call AI for analysis
        from ai_integration.services import USPAnalyzer
        analyzer = USPAnalyzer()
        analysis_result = analyzer.analyze_for_usps(
            website_content=website_content,
            usp_templates=templates_formatted
        )

        from datetime import datetime
        return JsonResponse({
            'success': True,
            'analysis': analysis_result,
            'scraped_at': datetime.now().isoformat(),
            'scraped_content_length': len(website_content),
            'pages_scraped': scrape_result.get('pages_scraped', 1),
            'total_urls_found': scrape_result.get('total_urls_found', 1)
        })

    except ValueError as e:
        # API key not configured
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
def scrape_and_detect_services_ajax(request):
    """
    AJAX endpoint til at scrape hjemmeside og auto-detektere services.
    Kaldes fra Step 1 inden brugeren går videre til Step 2.

    Accepts:
        - website_url: URL to scrape
        - scrape_mode: Number of pages (10, 50, 100, or 0 for all)

    Returns:
        - detected_services: List of detected service IDs with confidence
        - detected_industries: List of industry names
        - scraped_data: The scraped content for later use
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, json_dumps_params={'ensure_ascii': False})

    try:
        data = json.loads(request.body)
        website_url = data.get('website_url', '')
        scrape_mode = data.get('scrape_mode', 10)
        use_playwright = data.get('use_playwright', False)  # For JavaScript content (Trustpilot etc.)

        if not website_url:
            return JsonResponse({
                'success': True,
                'detected_services': [],
                'detected_industries': [],
                'scraped_at': None
            }, json_dumps_params={'ensure_ascii': False})

        # Step 1: Scrape website
        from ai_integration.services import ComprehensiveWebsiteScraper
        scraper = ComprehensiveWebsiteScraper()

        max_pages = scrape_mode if scrape_mode > 0 else None
        scrape_result = scraper.scrape_website(
            url=website_url,
            max_pages=max_pages,
            use_playwright=use_playwright
        )

        if not scrape_result or not scrape_result.get('combined_content'):
            return JsonResponse({
                'success': True,
                'detected_services': [],
                'detected_industries': [],
                'scraped_at': None,
                'message': 'Could not scrape website'
            }, json_dumps_params={'ensure_ascii': False})

        website_content = scrape_result['combined_content']

        # Also get service_summary which contains condensed info from ALL pages
        service_summary = scrape_result.get('service_summary', '')
        if service_summary:
            # Prepend the service summary so AI sees all page paths/titles first
            website_content = f"=== OVERSIGT OVER ALLE SIDER ===\n{service_summary}\n\n=== DETALJERET INDHOLD ===\n{website_content}"

        # Step 2: Get all available services
        all_services = []
        for service in IndustryService.objects.filter(is_active=True).select_related('industry'):
            all_services.append({
                'id': service.id,
                'name': service.name,
                'industry_id': service.industry_id,
                'industry_name': service.industry.name,
                'description': service.description or ''
            })

        if not all_services:
            return JsonResponse({
                'success': True,
                'detected_services': [],
                'detected_industries': [],
                'scraped_at': None,
                'message': 'No services configured'
            }, json_dumps_params={'ensure_ascii': False})

        # Step 3: Detect services using AI
        from ai_integration.services import ServiceDetector
        detector = ServiceDetector()
        detection_result = detector.detect_services(
            website_content=website_content,
            available_services=all_services
        )

        from datetime import datetime

        # Build response with service and industry IDs
        detected_service_ids = [
            svc['service_id'] for svc in detection_result.get('detected_services', [])
        ]

        # Map industry names to IDs (with fuzzy matching)
        detected_industry_ids = []
        if detection_result.get('detected_industries'):
            for industry_name in detection_result['detected_industries']:
                # Try exact match first
                industry = Industry.objects.filter(name__iexact=industry_name).first()

                # Try contains match (e.g., "El" matches "Elektriker")
                if not industry:
                    industry = Industry.objects.filter(name__icontains=industry_name).first()

                # Try if industry name contains the detected name
                if not industry:
                    for ind in Industry.objects.all():
                        if industry_name.lower() in ind.name.lower() or ind.name.lower() in industry_name.lower():
                            industry = ind
                            break

                if industry and industry.id not in detected_industry_ids:
                    detected_industry_ids.append(industry.id)

        # Also get industries from detected services (more reliable)
        for svc in detection_result.get('detected_services', []):
            service = IndustryService.objects.filter(id=svc['service_id']).select_related('industry').first()
            if service and service.industry_id not in detected_industry_ids:
                detected_industry_ids.append(service.industry_id)

        # Build scraped_pages dict with meta tags and sections for frontend
        scraped_pages = {}
        for path, page_info in scrape_result.get('pages', {}).items():
            scraped_pages[path] = {
                'url': page_info.get('url'),
                'path': path,
                'content': page_info.get('content', '')[:1000],  # Limit content for JSON response
                'meta_title': page_info.get('meta_title'),
                'meta_description': page_info.get('meta_description'),
                'page_type': page_info.get('page_type'),
                'sections': page_info.get('sections', []),  # Strukturerede sektioner fra siden
                'review_section_position': page_info.get('review_section_position'),  # Position af reviews på siden
            }

        return JsonResponse({
            'success': True,
            'detected_services': detection_result.get('detected_services', []),
            'detected_service_ids': detected_service_ids,
            'suggested_services': detection_result.get('suggested_services', []),
            'detected_industries': detection_result.get('detected_industries', []),
            'detected_industry_ids': detected_industry_ids,
            'primary_industry': detection_result.get('primary_industry'),
            'suggested_industry': detection_result.get('suggested_industry'),  # For industries not in DB
            'confidence_scores': detection_result.get('confidence_scores', {}),
            'scraped_at': datetime.now().isoformat(),
            'pages_scraped': scrape_result.get('pages_scraped', 1),
            'content_length': len(website_content),
            'scraped_pages': scraped_pages,  # Include pages with meta tags
            'extracted_reviews': scrape_result.get('extracted_reviews', []),  # Trustpilot/Google reviews
            'use_playwright': use_playwright,
        }, json_dumps_params={'ensure_ascii': False})

    except ValueError as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, json_dumps_params={'ensure_ascii': False})


def generate_seo_content_ajax(request):
    """
    AJAX endpoint til at generere eller omskrive SEO-indhold med AI.

    Accepts:
        - action: 'generate_new' eller 'rewrite'
        - service_name: Navn på servicen
        - industry: Branche (for nye sider)
        - usps: Liste af USP'er
        - company_name: Virksomhedsnavn (for nye sider)
        - company_profile: Virksomhedsprofil JSON (fra virksomhedsbeskrivelse-trinnet)
        - city: By/geografisk område (for lokal SEO)
        - existing_content: Eksisterende indhold (for omskrivning)
        - existing_meta_title: Nuværende meta titel (for omskrivning)
        - existing_meta_description: Nuværende meta beskrivelse (for omskrivning)

    Returns:
        - meta_title: Genereret meta titel
        - meta_description: Genereret meta beskrivelse
        - intro_text: Genereret intro-tekst (AIDA-model, 800 ord)
        - reviews: Array af detekterede kundeanmeldelser [{author, rating, text, platform, position}]
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, json_dumps_params={'ensure_ascii': False})

    try:
        data = json.loads(request.body)
        action = data.get('action', 'generate_new')
        service_name = data.get('service_name', '')
        industry = data.get('industry', '')
        usps = data.get('usps', [])
        company_name = data.get('company_name', '')
        company_profile = data.get('company_profile', None)  # JSON dict or None
        city = data.get('city', '')  # Geographic area for local SEO
        existing_content = data.get('existing_content', '')
        existing_meta_title = data.get('existing_meta_title', '')
        existing_meta_description = data.get('existing_meta_description', '')

        if not service_name:
            return JsonResponse({
                'success': False,
                'error': 'service_name is required'
            }, json_dumps_params={'ensure_ascii': False})

        # Generate content using AI
        from ai_integration.services import DescriptionGenerator
        generator = DescriptionGenerator()

        result = generator.generate_page_seo_content(
            action=action,
            service_name=service_name,
            industry=industry,
            usps=usps,
            company_name=company_name,
            company_profile=company_profile,
            city=city,
            existing_content=existing_content,
            existing_meta_title=existing_meta_title,
            existing_meta_description=existing_meta_description
        )

        return JsonResponse({
            'success': True,
            'meta_title': result.get('meta_title', ''),
            'meta_description': result.get('meta_description', ''),
            'intro_text': result.get('intro_text', ''),
            'reviews': result.get('reviews', [])
        }, json_dumps_params={'ensure_ascii': False})

    except ValueError as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, json_dumps_params={'ensure_ascii': False})


# ==========================================
# POSTAL CODE MANAGER VIEWS
# ==========================================

def postal_manager(request):
    """
    Postnummer Manager - administrer postnumre, visningsnavne og ekstra bynavne.
    Skjuler konsoliderede postnumre (København K, V, Frederiksberg C) undtagen primary.
    """
    from django.db.models import Q

    # Consolidated postal code ranges (hide all except primary)
    # København K: 1050-1473 (primary: 1050)
    # København V: 1550-1799 (primary: 1550)
    # Frederiksberg C: 1800-1999 (primary: 1800)
    postal_codes = PostalCode.objects.exclude(
        # Exclude København K (1051-1473) - keep 1050
        Q(code__gte='1051', code__lte='1473') |
        # Exclude København V (1551-1799) - keep 1550
        Q(code__gte='1551', code__lte='1799') |
        # Exclude Frederiksberg C (1801-1999) - keep 1800
        Q(code__gte='1801', code__lte='1999')
    ).order_by('code')

    return render(request, 'campaigns/postal_manager.html', {
        'postal_codes': postal_codes,
        'total_count': postal_codes.count()
    })


@csrf_exempt
def update_postal_code_ajax(request):
    """
    AJAX endpoint til at opdatere et postnummer's display_name eller additional_names.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    import json
    try:
        data = json.loads(request.body)
        code = data.get('code')
        display_name = data.get('display_name', '').strip()
        additional_names = data.get('additional_names', '').strip()

        if not code:
            return JsonResponse({'success': False, 'error': 'Missing postal code'})

        postal = get_object_or_404(PostalCode, code=code)

        # Opdater felter
        postal.display_name = display_name
        postal.additional_names = additional_names
        postal.save()

        return JsonResponse({
            'success': True,
            'code': postal.code,
            'display_name': postal.get_display_name(),
            'additional_names': postal.additional_names,
            'all_names': postal.get_all_names()
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def get_postal_codes_api(request):
    """
    API endpoint til at hente alle postnumre med customizations.
    Bruges af geo map til at matche bynavne.
    """
    postal_codes = PostalCode.objects.all().order_by('code')

    data = []
    for pc in postal_codes:
        data.append({
            'code': pc.code,
            'dawa_name': pc.dawa_name,
            'display_name': pc.get_display_name(),
            'additional_names': pc.additional_names,
            'all_names': pc.get_all_names()
        })

    return JsonResponse({'postal_codes': data})


@csrf_exempt
def export_campaign_builder_csv(request):
    """
    Eksporter Campaign Builder data til Google Ads Editor CSV format.
    Inkluderer kampagner, annoncegrupper, søgeord og negative søgeord.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST requests allowed'})

    try:
        import pandas as pd
        from io import BytesIO

        data = json.loads(request.body)
        campaigns_data = data.get('campaigns', {})

        if not campaigns_data:
            return JsonResponse({'success': False, 'error': 'Ingen kampagner at eksportere'})

        # Google Ads Editor kolonner (126 kolonner)
        columns = [
            'Campaign', 'Labels', 'Campaign Type', 'Networks', 'Budget', 'Budget type',
            'EU political ads', 'Standard conversion goals', 'Customer acquisition', 'Languages',
            'Bid Strategy Type', 'Bid Strategy Name', 'Enhanced CPC', 'Maximum CPC bid limit',
            'Start Date', 'End Date', 'Broad match keywords', 'Ad Schedule', 'Ad rotation',
            'Content exclusions', 'Targeting method', 'Exclusion method', 'Audience targeting',
            'Flexible Reach', 'AI Max', 'Text customization', 'Final URL expansion', 'Ad Group',
            'Max CPC', 'Max CPM', 'Target CPA', 'Max CPV', 'Target CPV', 'Percent CPC',
            'Target CPM', 'Target ROAS', 'Target CPC', 'Desktop Bid Modifier', 'Mobile Bid Modifier',
            'Tablet Bid Modifier', 'TV Screen Bid Modifier', 'Display Network Custom Bid Type',
            'Optimized targeting', 'Strict age and gender targeting', 'Search term matching',
            'Ad Group Type', 'Channels', 'Audience name', 'Age demographic', 'Gender demographic',
            'Income demographic', 'Parental status demographic', 'Remarketing audience segments',
            'Interest categories', 'Life events', 'Custom audience segments', 'Detailed demographics',
            'Remarketing audience exclusions', 'Tracking template', 'Final URL suffix',
            'Custom parameters', 'ID', 'Location', 'Reach', 'Location groups', 'Radius',
            'Unit', 'Bid Modifier', 'Keyword', 'Criterion Type', 'First page bid',
            'Top of page bid', 'First position bid', 'Quality score', 'Landing page experience',
            'Expected CTR', 'Ad relevance', 'Final URL', 'Final mobile URL', 'Ad type',
            'Headline 1', 'Headline 1 position', 'Headline 2', 'Headline 2 position',
            'Headline 3', 'Headline 3 position', 'Headline 4', 'Headline 4 position',
            'Headline 5', 'Headline 5 position', 'Headline 6', 'Headline 6 position',
            'Headline 7', 'Headline 7 position', 'Headline 8', 'Headline 8 position',
            'Headline 9', 'Headline 9 position', 'Headline 10', 'Headline 10 position',
            'Headline 11', 'Headline 11 position', 'Headline 12', 'Headline 12 position',
            'Headline 13', 'Headline 13 position', 'Headline 14', 'Headline 14 position',
            'Headline 15', 'Headline 15 position', 'Description 1', 'Description 1 position',
            'Description 2', 'Description 2 position', 'Description 3', 'Description 3 position',
            'Description 4', 'Description 4 position', 'Path 1', 'Path 2', 'Campaign Status',
            'Ad Group Status', 'Status', 'Approval Status', 'Ad strength', 'Comment'
        ]

        all_rows = []

        # Hent alle negative keyword lister på forhånd
        all_negative_lists = {}
        for nk_list in NegativeKeywordList.objects.prefetch_related('negative_keywords').all():
            all_negative_lists[str(nk_list.id)] = {
                'name': nk_list.name,
                'keywords': list(nk_list.negative_keywords.values_list('keyword_text', 'match_type'))
            }

        for industry_id, campaign_data in campaigns_data.items():
            campaign_name = campaign_data.get('industry_name', 'Campaign')
            # Understøt både 'budget' og 'daily_budget' felter
            budget = campaign_data.get('daily_budget') or campaign_data.get('budget', 500)
            if not budget:
                budget = 500
            negative_list_ids = campaign_data.get('negative_keyword_list_ids', [])
            ad_groups = campaign_data.get('ad_groups', {})

            # 1. Campaign row (matcher præcis den fungerende fil)
            campaign_row = {col: '' for col in columns}
            campaign_row.update({
                'Campaign': campaign_name,
                'Campaign Type': 'Search',
                'Networks': 'Google search',
                'Budget': f"{float(budget):.2f}",
                'Budget type': 'Daily',
                'EU political ads': "Doesn't have EU political ads",
                'Standard conversion goals': 'Account-level',
                'Customer acquisition': 'Bid equally',
                'Languages': 'da',
                'Bid Strategy Type': 'Maximize clicks',
                'Enhanced CPC': 'Disabled',
                'Maximum CPC bid limit': '35.00',
                'Start Date': '[]',
                'End Date': '[]',
                'Broad match keywords': 'Off',
                'Ad Schedule': '[]',
                'Ad rotation': 'Rotate evenly',
                'Content exclusions': '[]',
                'Targeting method': 'Location of presence or Area of interest',
                'Exclusion method': 'Location of presence',
                'Audience targeting': 'Audience segments',
                'Flexible Reach': 'Audience segments',
                'AI Max': 'Disabled',
                'Text customization': 'Disabled',
                'Final URL expansion': 'Disabled',
                'Campaign Status': 'Enabled'
            })
            all_rows.append(campaign_row)

            # 2. Ad Groups, Keywords, and Ads (først, før negative keywords)
            for service_id, ad_group_data in ad_groups.items():
                service_name = ad_group_data.get('service_name', 'Ad Group')
                ad_group_full_name = f"{campaign_name} - {service_name}"
                keywords = ad_group_data.get('keywords', [])
                headlines = ad_group_data.get('headlines', [])
                descriptions = ad_group_data.get('descriptions', [])

                # Ad Group row (matcher præcis den fungerende fil - INGEN Max CPC her)
                adgroup_row = {col: '' for col in columns}
                adgroup_row.update({
                    'Campaign': campaign_name,
                    'Languages': 'All',
                    'Audience targeting': 'Audience segments',
                    'Flexible Reach': 'Audience segments;Genders;Ages;Parental status;Household incomes',
                    'Ad Group': ad_group_full_name,
                    'Optimized targeting': 'Disabled',
                    'Strict age and gender targeting': 'Disabled',
                    'Search term matching': 'Enabled',
                    'Ad Group Type': 'Standard',
                    'Channels': '[]',
                    'Campaign Status': 'Enabled',
                    'Ad Group Status': 'Enabled'
                })
                all_rows.append(adgroup_row)

                # Keyword rows (matcher præcis den fungerende fil)
                for kw in keywords:
                    if isinstance(kw, dict):
                        # Understøt både 'keyword' og 'keyword_text' felter
                        kw_text = kw.get('keyword') or kw.get('keyword_text', '')
                        match_type = kw.get('match_type', 'Phrase')
                        final_url = kw.get('final_url') or kw.get('url', '')
                    else:
                        kw_text = str(kw)
                        match_type = 'Phrase'
                        final_url = ''

                    if not kw_text:
                        continue

                    keyword_row = {col: '' for col in columns}
                    keyword_row.update({
                        'Campaign': campaign_name,
                        'Ad Group': ad_group_full_name,
                        'Max CPC': '35.00',
                        'Keyword': kw_text,
                        'Criterion Type': match_type.title(),
                        'First page bid': '0.00',
                        'Top of page bid': '0.00',
                        'First position bid': '0.00',
                        'Landing page experience': ' -',
                        'Expected CTR': ' -',
                        'Ad relevance': ' -',
                        'Final URL': final_url,
                        'Campaign Status': 'Enabled',
                        'Ad Group Status': 'Enabled',
                        'Status': 'Enabled',
                        'Approval Status': 'Pending review'
                    })
                    all_rows.append(keyword_row)

                # Ad row (RSA - Responsive Search Ad) - matcher præcis den fungerende fil
                if headlines or descriptions:
                    # Brug første keyword's final_url eller fallback
                    first_url = ''
                    if keywords:
                        first_kw = keywords[0]
                        if isinstance(first_kw, dict):
                            first_url = first_kw.get('final_url') or first_kw.get('url', '')

                    ad_row = {col: '' for col in columns}
                    ad_row.update({
                        'Campaign': campaign_name,
                        'Ad Group': ad_group_full_name,
                        'Final URL': first_url,
                        'Ad type': 'Responsive search ad',
                        'Campaign Status': 'Enabled',
                        'Ad Group Status': 'Enabled',
                        'Status': 'Enabled',
                        'Approval Status': 'Pending review'
                    })

                    # Headlines (max 15)
                    for i, headline in enumerate(headlines[:15], 1):
                        if headline:
                            # Understøt både string og dict format
                            if isinstance(headline, dict):
                                headline_text = headline.get('text', '') or headline.get('headline', '') or str(headline)
                            else:
                                headline_text = str(headline)
                            if headline_text:
                                ad_row[f'Headline {i}'] = headline_text[:30]

                    # Descriptions (max 4)
                    for i, description in enumerate(descriptions[:4], 1):
                        if description:
                            # Understøt både string og dict format
                            if isinstance(description, dict):
                                desc_text = description.get('text', '') or description.get('description', '') or str(description)
                            else:
                                desc_text = str(description)
                            if desc_text:
                                ad_row[f'Description {i}'] = desc_text[:90]

                    all_rows.append(ad_row)

            # 3. Negative Keywords for campaign (efter alle andre rækker)
            for list_id in negative_list_ids:
                list_id_str = str(list_id)
                if list_id_str in all_negative_lists:
                    neg_list = all_negative_lists[list_id_str]
                    for keyword_text, match_type in neg_list['keywords']:
                        neg_row = {col: '' for col in columns}
                        neg_row.update({
                            'Campaign': campaign_name,
                            'Keyword': keyword_text,
                            'Criterion Type': f"Negative {match_type.lower()}" if match_type else 'Negative phrase',
                            'Campaign Status': 'Enabled',
                            'Status': 'Enabled'
                        })
                        all_rows.append(neg_row)

        # Håndter GEO kampagner (geo_campaigns)
        geo_campaigns_data = data.get('geo_campaigns', {})
        for service_id, geo_campaign in geo_campaigns_data.items():
            geo_campaign_name = geo_campaign.get('name', f'GEO Kampagne {service_id}')
            geo_budget = geo_campaign.get('daily_budget') or 500
            geo_neg_list_ids = geo_campaign.get('negative_keyword_list_ids', [])
            geo_ad_group = geo_campaign.get('ad_group', {})

            # GEO Campaign row
            geo_campaign_row = {col: '' for col in columns}
            geo_campaign_row.update({
                'Campaign': geo_campaign_name,
                'Campaign Type': 'Search',
                'Networks': 'Google search',
                'Budget': f"{float(geo_budget):.2f}",
                'Budget type': 'Daily',
                'EU political ads': "Doesn't have EU political ads",
                'Standard conversion goals': 'Account-level',
                'Customer acquisition': 'Bid equally',
                'Languages': 'da',
                'Bid Strategy Type': 'Maximize clicks',
                'Enhanced CPC': 'Disabled',
                'Maximum CPC bid limit': '35.00',
                'Start Date': '[]',
                'End Date': '[]',
                'Broad match keywords': 'Off',
                'Ad Schedule': '[]',
                'Ad rotation': 'Rotate evenly',
                'Content exclusions': '[]',
                'Targeting method': 'Location of presence or Area of interest',
                'Exclusion method': 'Location of presence',
                'Audience targeting': 'Audience segments',
                'Flexible Reach': 'Audience segments',
                'AI Max': 'Disabled',
                'Text customization': 'Disabled',
                'Final URL expansion': 'Disabled',
                'Campaign Status': 'Enabled'
            })
            all_rows.append(geo_campaign_row)

            # GEO Ad Group row
            geo_ad_group_name = f"{geo_campaign_name} - Hovedgruppe"
            geo_adgroup_row = {col: '' for col in columns}
            geo_adgroup_row.update({
                'Campaign': geo_campaign_name,
                'Languages': 'All',
                'Audience targeting': 'Audience segments',
                'Flexible Reach': 'Audience segments;Genders;Ages;Parental status;Household incomes',
                'Ad Group': geo_ad_group_name,
                'Optimized targeting': 'Disabled',
                'Strict age and gender targeting': 'Disabled',
                'Search term matching': 'Enabled',
                'Ad Group Type': 'Standard',
                'Channels': '[]',
                'Campaign Status': 'Enabled',
                'Ad Group Status': 'Enabled'
            })
            all_rows.append(geo_adgroup_row)

            # GEO Keywords
            geo_keywords = geo_ad_group.get('keywords', [])
            for kw in geo_keywords:
                if isinstance(kw, dict):
                    kw_text = kw.get('keyword', '')
                    match_type = kw.get('match_type', 'Phrase')
                    final_url = kw.get('url', '')
                else:
                    continue

                if not kw_text:
                    continue

                geo_keyword_row = {col: '' for col in columns}
                geo_keyword_row.update({
                    'Campaign': geo_campaign_name,
                    'Ad Group': geo_ad_group_name,
                    'Max CPC': '35.00',
                    'Keyword': kw_text,
                    'Criterion Type': match_type.title(),
                    'First page bid': '0.00',
                    'Top of page bid': '0.00',
                    'First position bid': '0.00',
                    'Landing page experience': ' -',
                    'Expected CTR': ' -',
                    'Ad relevance': ' -',
                    'Final URL': final_url,
                    'Campaign Status': 'Enabled',
                    'Ad Group Status': 'Enabled',
                    'Status': 'Enabled',
                    'Approval Status': 'Pending review'
                })
                all_rows.append(geo_keyword_row)

            # GEO Ad row
            geo_headlines = geo_ad_group.get('headlines', [])
            geo_descriptions = geo_ad_group.get('descriptions', [])
            if geo_headlines or geo_descriptions:
                first_geo_url = ''
                if geo_keywords:
                    first_geo_kw = geo_keywords[0]
                    if isinstance(first_geo_kw, dict):
                        first_geo_url = first_geo_kw.get('url', '')

                geo_ad_row = {col: '' for col in columns}
                geo_ad_row.update({
                    'Campaign': geo_campaign_name,
                    'Ad Group': geo_ad_group_name,
                    'Final URL': first_geo_url,
                    'Ad type': 'Responsive search ad',
                    'Campaign Status': 'Enabled',
                    'Ad Group Status': 'Enabled',
                    'Status': 'Enabled',
                    'Approval Status': 'Pending review'
                })

                for i, headline in enumerate(geo_headlines[:15], 1):
                    if headline:
                        # Understøt både string og dict format
                        if isinstance(headline, dict):
                            headline_text = headline.get('text', '') or headline.get('headline', '') or str(headline)
                        else:
                            headline_text = str(headline)
                        if headline_text:
                            geo_ad_row[f'Headline {i}'] = headline_text[:30]

                for i, description in enumerate(geo_descriptions[:4], 1):
                    if description:
                        # Understøt både string og dict format
                        if isinstance(description, dict):
                            desc_text = description.get('text', '') or description.get('description', '') or str(description)
                        else:
                            desc_text = str(description)
                        if desc_text:
                            geo_ad_row[f'Description {i}'] = desc_text[:90]

                all_rows.append(geo_ad_row)

            # GEO Negative Keywords
            for list_id in geo_neg_list_ids:
                list_id_str = str(list_id)
                if list_id_str in all_negative_lists:
                    neg_list = all_negative_lists[list_id_str]
                    for keyword_text, match_type in neg_list['keywords']:
                        neg_row = {col: '' for col in columns}
                        neg_row.update({
                            'Campaign': geo_campaign_name,
                            'Keyword': keyword_text,
                            'Criterion Type': f"Negative {match_type.lower()}" if match_type else 'Negative phrase',
                            'Campaign Status': 'Enabled',
                            'Status': 'Enabled'
                        })
                        all_rows.append(neg_row)

        # Opret DataFrame og eksporter
        df = pd.DataFrame(all_rows)

        # Ensure all columns exist
        for col in columns:
            if col not in df.columns:
                df[col] = ''

        # Reorder columns
        df = df[columns]

        # Opret CSV response med korrekt encoding (UTF-16LE med tab separator for Google Ads Editor)
        response = HttpResponse(content_type='text/csv; charset=utf-16le')
        filename = f"Google_Ads_Export_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Skriv CSV med UTF-16 encoding og tab separator (inkluderer BOM automatisk)
        csv_content = df.to_csv(index=False, sep='\t', encoding='utf-16le', lineterminator='\r\n')
        response.write(csv_content)

        return response

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        import traceback
        return JsonResponse({'success': False, 'error': str(e), 'traceback': traceback.format_exc()})


# =====================================================
# Programmatic Byside AJAX Endpoints
# =====================================================

@csrf_exempt
def crawl_sitemap_ajax(request):
    """
    AJAX endpoint til sitemap crawling.
    POST: {website_url: str}
    Returns: {success: bool, urls_found: int, urls: List[str], message: str}
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        from .sitemap_service import SitemapCrawler

        data = json.loads(request.body)
        website_url = data.get('website_url', '').strip()

        if not website_url:
            return JsonResponse({'success': False, 'error': 'Website URL er påkrævet'})

        # Tilføj https hvis mangler
        if not website_url.startswith('http'):
            website_url = f'https://{website_url}'

        crawler = SitemapCrawler(website_url)
        success, urls, message = crawler.crawl_all_urls()

        return JsonResponse({
            'success': success,
            'urls_found': len(urls),
            'urls': urls[:1000],  # Begræns til 1000 URLs for performance
            'message': message
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
def match_city_pages_ajax(request):
    """
    AJAX endpoint til URL-matching mod sitemap.
    POST: {
        website_url: str,
        service_name: str,
        cities: List[str]
    }
    Returns: {
        success: bool,
        results: [{city, status, existing_url}],
        existing_count: int,
        missing_count: int
    }
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        from .sitemap_service import crawl_and_match

        data = json.loads(request.body)
        website_url = data.get('website_url', '').strip()
        service_name = data.get('service_name', '').strip()
        cities = data.get('cities', [])

        if not website_url:
            return JsonResponse({'success': False, 'error': 'Website URL er påkrævet'})
        if not service_name:
            return JsonResponse({'success': False, 'error': 'Service navn er påkrævet'})
        if not cities:
            return JsonResponse({'success': False, 'error': 'Mindst én by skal vælges'})

        # Tilføj https hvis mangler
        if not website_url.startswith('http'):
            website_url = f'https://{website_url}'

        result = crawl_and_match(website_url, service_name, cities)

        return JsonResponse(result)

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
def generate_programmatic_descriptions_ajax(request):
    """
    AJAX endpoint til AI-generering af descriptions og meta tags med {BYNAVN} placeholder.
    POST: {
        service_name: str,
        industry_name: str,
        usps: List[str],
        keywords: List[str],
        generate_meta_tags: bool (optional) - hvis true, generér 7 meta titler + 7 meta beskrivelser
    }
    Returns:
        - Uden generate_meta_tags: {success: bool, descriptions: List[str]}
        - Med generate_meta_tags: {success: bool, meta_titles: List[str], meta_descriptions: List[str]}
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        from ai_integration.services import DescriptionGenerator
        from .models import ServiceMetaExample

        data = json.loads(request.body)
        service_name = data.get('service_name', '')
        industry_name = data.get('industry_name', '')
        usps = data.get('usps', [])
        keywords = data.get('keywords', [])
        generate_meta_tags = data.get('generate_meta_tags', False)
        service_id = data.get('service_id')  # NEW: Optional service ID for few-shot examples

        if not service_name:
            return JsonResponse({'success': False, 'error': 'Service navn er påkrævet'})

        # Initialiser generator
        generator = DescriptionGenerator()

        # Hvis generate_meta_tags er true, generér 7 meta titler + 7 meta beskrivelser
        if generate_meta_tags:
            # Fetch few-shot examples if service_id is provided
            few_shot_examples = None
            if service_id:
                try:
                    # Get random 10 examples from the service
                    examples = ServiceMetaExample.objects.filter(
                        service_id=service_id
                    ).order_by('?')[:10]

                    if examples.exists():
                        few_shot_examples = [
                            {
                                'meta_title': ex.meta_title,
                                'meta_description': ex.meta_description
                            }
                            for ex in examples
                        ]
                        print(f"Using {len(few_shot_examples)} few-shot examples for service {service_id}")
                except Exception as e:
                    print(f"Warning: Could not fetch few-shot examples: {e}")

            result = generator.generate_meta_tags(
                service_name=service_name,
                usps=usps,
                few_shot_examples=few_shot_examples
            )
            return JsonResponse({
                'success': True,
                'meta_titles': result['meta_titles'],
                'meta_descriptions': result['meta_descriptions']
            })

        # Ellers generér standard descriptions
        # Tilføj instruktion om {BYNAVN} placeholder til prompt
        geo_usps = usps + [
            "VIGTIGT: Inkluder {BYNAVN} placeholder i mindst 2 af beskrivelserne",
            "Eksempel: 'Professionel service i {BYNAVN}. Ring nu!'"
        ]

        descriptions = generator.generate_descriptions(
            service_name=service_name,
            industry_name=industry_name,
            usps=geo_usps,
            keywords=keywords
        )

        return JsonResponse({
            'success': True,
            'descriptions': descriptions
        })

    except ValueError as e:
        # API key ikke konfigureret
        return JsonResponse({'success': False, 'error': str(e)})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
def generate_seo_meta_ajax(request):
    """Generate SEO meta title and description using AI with few-shot examples."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        data = json.loads(request.body)
        service_name = data.get('service_name', '')
        service_id = data.get('service_id')  # Optional: for few-shot examples
        seo_keywords = data.get('seo_keywords', [])
        usps = data.get('usps', [])

        # Debug: Log received data
        print(f"SEO Meta Request: service_name='{service_name}', service_id={service_id}, type={type(service_id)}")

        if not service_name:
            return JsonResponse({'success': False, 'error': 'Service name required'})

        # Fetch few-shot examples if service_id is provided
        from campaigns.models import ServiceMetaExample
        few_shot_examples = None
        if service_id:
            try:
                examples = ServiceMetaExample.objects.filter(
                    service_id=service_id
                ).order_by('?')[:5]

                if examples.exists():
                    few_shot_examples = [
                        {
                            'meta_title': ex.meta_title,
                            'meta_description': ex.meta_description
                        }
                        for ex in examples
                    ]
                    print(f"SEO Meta: Using {len(few_shot_examples)} few-shot examples for service {service_id}")
            except Exception as e:
                print(f"Warning: Could not fetch few-shot examples: {e}")

        # Initialize generator
        from ai_integration.services import DescriptionGenerator
        generator = DescriptionGenerator()

        # Generate SEO meta tags (WITHOUT {BYNAVN} requirement)
        result = generator.generate_seo_meta_tags(
            service_name=service_name,
            usps=usps,
            seo_keywords=seo_keywords,
            few_shot_examples=few_shot_examples
        )

        return JsonResponse({
            'success': True,
            'meta_title': result['meta_title'],
            'meta_description': result['meta_description']
        })

    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
def export_seo_content_csv(request):
    """
    Export SEO content to WordPress-compatible CSV (WP All Import).

    Returns CSV with columns:
    - service_name: Name of the service
    - page_path: URL path (e.g., /elektriker/)
    - meta_title: Meta title for the page
    - meta_description: Meta description
    - sections_json: JSON array of all sections [{header, content, tag}, ...]
    - full_content: HTML-formatted content for simple WordPress import
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        import csv
        from io import StringIO

        data = json.loads(request.body)
        seo_pages = data.get('seo_pages', {})
        website_url = data.get('website_url', '')

        if not seo_pages:
            return JsonResponse({'success': False, 'error': 'No SEO pages to export'})

        output = StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)

        # Header row - WordPress WP All Import compatible
        writer.writerow([
            'service_name',
            'page_path',
            'full_url',
            'meta_title',
            'meta_description',
            'sections_json',
            'full_content'
        ])

        # Sort pages by service name for consistent output
        sorted_pages = sorted(seo_pages.items(), key=lambda x: x[1].get('service_name', ''))

        for service_id, page_data in sorted_pages:
            service_name = page_data.get('service_name', '')
            source_path = page_data.get('source_path', '/')
            meta_title = page_data.get('meta_title', '')
            meta_description = page_data.get('meta_description', '')
            sections = page_data.get('sections', [])

            # Build full URL
            base_url = website_url.rstrip('/')
            full_url = f"{base_url}{source_path}" if base_url else source_path

            # Build full_content as HTML
            html_parts = []
            for section in sections:
                tag = section.get('tag', 'h2')
                header = section.get('header', '')
                content = section.get('content', '')

                if header:
                    html_parts.append(f'<{tag}>{header}</{tag}>')
                if content:
                    # Split content into paragraphs
                    for para in content.split('\n\n'):
                        para = para.strip()
                        if para:
                            html_parts.append(f'<p>{para}</p>')

            full_content = '\n'.join(html_parts)

            # Write row
            writer.writerow([
                service_name,
                source_path,
                full_url,
                meta_title,
                meta_description,
                json.dumps(sections, ensure_ascii=False),
                full_content
            ])

        # Create response
        csv_content = output.getvalue()

        # Add BOM for Excel UTF-8 compatibility
        response = HttpResponse(
            '\ufeff' + csv_content,
            content_type='text/csv; charset=utf-8'
        )
        response['Content-Disposition'] = 'attachment; filename="seo_content_export.csv"'

        return response

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# =====================================================
# CLIENT MANAGEMENT VIEWS
# =====================================================

def client_list(request):
    """Listevisning over kunder."""
    clients = Client.objects.all().select_related('industry')

    # Add campaign count to each client
    for client in clients:
        client.campaign_count = Campaign.objects.filter(client=client).count()

    industries = Industry.objects.all()

    return render(request, 'campaigns/client_list.html', {
        'clients': clients,
        'industries': industries
    })


def client_detail(request, client_id):
    """Kundebillede - fuld profil med AI-data og segmenteret overblik."""
    from .models import GeographicRegion, DanishCity, IndustryService

    client = get_object_or_404(Client, id=client_id)
    campaigns = Campaign.objects.filter(client=client).order_by('-created_at')

    # Get all geographic regions with cities for the map
    geographic_regions = GeographicRegion.objects.filter(is_active=True).prefetch_related('cities').order_by('name')

    # Parse client's campaign_config if exists
    campaign_config = client.campaign_config or {}
    company_info = campaign_config.get('company_info', {})
    geo_config = campaign_config.get('geo_config', {})
    campaigns_config = campaign_config.get('campaigns', {})

    # Get selected region IDs from client
    selected_region_ids = client.geographic_regions or []

    # Get industries and services from campaign_config
    selected_industry_ids = campaign_config.get('industry_ids', [])
    selected_service_ids = campaign_config.get('service_ids', [])

    # Fetch actual industry and service objects
    industries_with_services = []
    if selected_industry_ids:
        for industry in Industry.objects.filter(id__in=selected_industry_ids):
            services = IndustryService.objects.filter(
                industry=industry,
                id__in=selected_service_ids
            ) if selected_service_ids else []
            industries_with_services.append({
                'industry': industry,
                'services': list(services)
            })

    # Get all USPs (both predefined and AI-detected)
    # Replace variables with actual values from campaign_config
    predefined_usps = []
    usp_ids = campaign_config.get('usp_ids', [])
    usp_variable_values = campaign_config.get('usp_variable_values', {})
    usp_custom_texts = campaign_config.get('usp_custom_texts', {})

    if usp_ids:
        from usps.models import USPTemplate
        import re

        for usp in USPTemplate.objects.filter(id__in=usp_ids):
            usp_id_str = str(usp.id)

            # Check if there's a custom text (fully edited)
            if usp_id_str in usp_custom_texts:
                display_text = usp_custom_texts[usp_id_str]
            else:
                # Replace variables with user values
                display_text = usp.text
                user_values = usp_variable_values.get(usp_id_str, {})

                # Find all variables like {VAR1:default/options} and replace them
                def replace_var(match):
                    full_match = match.group(0)
                    var_name = match.group(1)  # e.g., "VAR1"
                    var_index = var_name.replace('VAR', '')  # e.g., "1"

                    # Get user value or extract default from the variable definition
                    if var_index in user_values:
                        return str(user_values[var_index])

                    # Extract default value from pattern like {VAR1:default/option1/option2}
                    var_content = match.group(2) if match.lastindex >= 2 else ''
                    if var_content:
                        # Default is everything before the first /
                        default_value = var_content.split('/')[0] if '/' in var_content else var_content
                        return default_value

                    return full_match  # Keep original if no replacement found

                # Pattern matches {VAR1:content} or {VAR1}
                display_text = re.sub(r'\{(VAR\d+)(?::([^}]*))?\}', replace_var, display_text)

            # Create a simple object with the display text
            predefined_usps.append({
                'id': usp.id,
                'text': display_text,
                'main_category': usp.main_category.name if usp.main_category else None
            })

    custom_usps = campaign_config.get('custom_usps', [])
    detected_usps = client.detected_usps or []

    # Get crawled pages from crawlState
    crawl_state = campaign_config.get('crawlState', {})
    scraped_pages = crawl_state.get('scraped_pages', {})
    extracted_reviews = crawl_state.get('extracted_reviews', [])

    # Build cities comparison data
    # Get all cities from selected regions
    selected_cities = set()
    if selected_region_ids:
        for region in geographic_regions:
            if region.id in selected_region_ids:
                for city in region.cities.all():
                    selected_cities.add(city.city_name)

    # Get cities that already have campaigns/bysider created
    created_bysider = set()
    byside_urls = geo_config.get('byside_urls', [])
    for byside in byside_urls:
        if byside.get('exists') or byside.get('edited'):
            created_bysider.add(byside.get('city', ''))

    # Format company profile as pretty JSON for display
    if company_info.get('profile'):
        company_info['profile_json'] = json.dumps(company_info['profile'], indent=2, ensure_ascii=False)

    return render(request, 'campaigns/client_detail.html', {
        'client': client,
        'campaigns': campaigns,
        'geographic_regions': geographic_regions,
        'selected_region_ids': selected_region_ids,
        'campaign_config': campaign_config,
        'company_info': company_info,
        'geo_config': geo_config,
        'industries_with_services': industries_with_services,
        'predefined_usps': predefined_usps,
        'custom_usps': custom_usps,
        'detected_usps': detected_usps,
        'selected_cities': list(selected_cities),
        'created_bysider': list(created_bysider),
        'byside_urls': byside_urls,
        'scraped_pages': scraped_pages,
        'extracted_reviews': extracted_reviews,
        'crawl_state': crawl_state,
        'google_maps_api_key': 'AIzaSyBDH6MTS0Hq0ISb0bNQjEAC14321pzM0jw',
    })


@csrf_exempt
def get_client_ajax(request, client_id):
    """Hent enkelt kunde data via AJAX."""
    try:
        client = get_object_or_404(Client, id=client_id)
        return JsonResponse({
            'success': True,
            'client': {
                'id': client.id,
                'name': client.name,
                'website_url': client.website_url,
                'industry_id': client.industry_id,
                'description': client.description or ''
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
def create_client_ajax(request):
    """Opret ny kunde via AJAX."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        data = json.loads(request.body)

        industry = None
        if data.get('industry_id'):
            industry = get_object_or_404(Industry, id=data['industry_id'])

        # Get user if authenticated, otherwise None
        created_by = request.user if request.user.is_authenticated else None

        client = Client.objects.create(
            name=data['name'],
            website_url=data['website_url'],
            industry=industry,
            description=data.get('description', ''),
            created_by=created_by
        )

        return JsonResponse({
            'success': True,
            'client_id': client.id,
            'message': 'Kunde oprettet'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
def update_client_ajax(request, client_id):
    """Opdater kunde via AJAX."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        client = get_object_or_404(Client, id=client_id)
        data = json.loads(request.body)

        client.name = data.get('name', client.name)
        client.website_url = data.get('website_url', client.website_url)
        client.description = data.get('description', client.description)

        if data.get('industry_id'):
            client.industry = get_object_or_404(Industry, id=data['industry_id'])
        elif 'industry_id' in data and data['industry_id'] is None:
            client.industry = None

        client.save()

        return JsonResponse({
            'success': True,
            'message': 'Kunde opdateret'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
def delete_client_ajax(request, client_id):
    """Slet kunde via AJAX."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        client = get_object_or_404(Client, id=client_id)
        client.delete()

        return JsonResponse({
            'success': True,
            'message': 'Kunde slettet'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
def save_client_from_builder(request):
    """Gem Campaign Builder data som kunde."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        data = json.loads(request.body)

        if data.get('client_id'):
            # Opdater eksisterende kunde
            client = get_object_or_404(Client, id=data['client_id'])
        else:
            # Opret ny kunde
            if not data.get('name'):
                return JsonResponse({'success': False, 'error': 'Kundenavn er påkrævet'})
            client = Client(name=data['name'])

        # Opdater felter
        if data.get('website_url'):
            client.website_url = data['website_url']
        if data.get('description'):
            client.description = data['description']
        if data.get('company_profile'):
            client.company_profile = data['company_profile']
        if data.get('detected_services'):
            client.detected_services = data['detected_services']
        if data.get('detected_usps'):
            client.detected_usps = data['detected_usps']
        if data.get('selected_purposes') is not None:
            client.selected_purposes = data['selected_purposes']
        if data.get('geographic_regions') is not None:
            client.geographic_regions = data['geographic_regions']
        if data.get('selected_services') is not None:
            client.selected_services = data['selected_services']
        if data.get('selected_usps') is not None:
            client.selected_usps = data['selected_usps']
        if data.get('campaign_config'):
            client.campaign_config = data['campaign_config']

        client.save()

        return JsonResponse({
            'success': True,
            'client_id': client.id,
            'message': 'Kunde gemt fra Campaign Builder'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def get_client_campaign_data(request, client_id):
    """Hent kunde data til Campaign Builder."""
    try:
        client = get_object_or_404(Client, id=client_id)

        return JsonResponse({
            'success': True,
            'client_id': client.id,
            'name': client.name,
            'website_url': client.website_url or '',
            'description': client.description or '',
            'company_profile': client.company_profile,
            'detected_services': client.detected_services,
            'detected_usps': client.detected_usps,
            'selected_purposes': client.selected_purposes or [],
            'geographic_regions': client.geographic_regions or [],
            'selected_services': client.selected_services or [],
            'selected_usps': client.selected_usps or [],
            'campaign_config': client.campaign_config,
            'scraped_data': client.scraped_data
        }, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
